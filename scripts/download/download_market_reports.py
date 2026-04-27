from __future__ import annotations

import argparse
from contextlib import suppress
from html.parser import HTMLParser
import logging
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import date
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlparse

try:
    import httpx
except ImportError:  # pragma: no cover - dependency availability varies per environment.
    httpx = None  # type: ignore[assignment]

try:
    from playwright.sync_api import Error as PlaywrightError
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
except ImportError:  # pragma: no cover - dependency availability varies per environment.
    PlaywrightError = Exception  # type: ignore[assignment]
    PlaywrightTimeoutError = TimeoutError  # type: ignore[assignment]
    sync_playwright = None  # type: ignore[assignment]

from openpyxl import load_workbook


LOGGER = logging.getLogger("download_market_reports")
EXPECTED_COMPANY_COUNT = 126

SECTOR_FOLDER = {
    "Reasuransi": "reinsurance",
    "Asuransi Umum": "general",
    "Asuransi Jiwa": "life",
}

FALLBACK_SHEET_GROUPS = (
    ("Reasuransi", 0, 1),
    ("Asuransi Umum", 3, 4),
    ("Asuransi Jiwa", 6, 7),
)

INVALID_SOURCE_VALUES = {"", "-", "n/a", "na", "eror", "error", "tidak beroperasi"}
BLOCKED_HTTP_STATUS_CODES = {401, 403, 406, 429}

PDF_HINTS = (
    "laporan",
    "keuangan",
    "financial",
    "monthly",
    "bulan",
    "statement",
    "publikasi",
    "report",
    "download",
    "unduh",
    "newspaper",
)
MONTHLY_HINTS = ("bulanan", "bulan", "monthly")
EXCLUDE_HINTS = (
    "annual",
    "tahunan",
    "keberlanjutan",
    "sustainability",
    "triwulan",
    "quarterly",
    "semester",
    "interim",
    "pengaduan",
    "ringkasan",
)

MONTH_ID = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}

MONTH_EN = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


class MissingDependencyError(RuntimeError):
    """Raised when an optional runtime dependency is required but unavailable."""


class SourceBlockedError(RuntimeError):
    """Raised when a site blocks the plain HTTP client and needs browser fallback."""


@dataclass(frozen=True)
class SourceEndpoint:
    label: str
    url: str
    kind: str


@dataclass(frozen=True)
class CompanyRecord:
    company_id: str
    company_name: str
    company_type: str
    sector_label: str
    primary_source: SourceEndpoint | None
    fallback_source: SourceEndpoint | None

    def ordered_sources(self) -> list[SourceEndpoint]:
        sources: list[SourceEndpoint] = []
        for source in (self.primary_source, self.fallback_source):
            if not source:
                continue
            if any(existing.url == source.url for existing in sources):
                continue
            sources.append(source)
        return sources


@dataclass(frozen=True)
class DiscoveredDocument:
    company_id: str
    company_name: str
    company_type: str
    source_url: str
    source_label: str
    source_kind: str
    discovery_mode: str
    document_url: str
    anchor_text: str | None
    report_year: int | None = None
    report_month: int | None = None
    report_period_label: str | None = None
    notes: str | None = None


class LinkExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._anchor_href: str | None = None
        self._anchor_text: list[str] = []
        self._anchor_attrs: dict[str, str] = {}
        self._script_tag: str | None = None
        self._script_buffer: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_map = {name.lower(): (value or "") for name, value in attrs}
        tag_lower = tag.lower()

        if tag_lower == "a" and attrs_map.get("href"):
            self._anchor_href = attrs_map["href"]
            self._anchor_text = []
            self._anchor_attrs = attrs_map

        if tag_lower in {"iframe", "embed", "source", "object"}:
            for attr in ("src", "data", "href"):
                if attrs_map.get(attr):
                    text = attrs_map.get("title") or attrs_map.get("aria-label") or attrs_map.get("alt") or ""
                    self.links.append((attrs_map[attr], text))

        for attr_name, attr_value in attrs_map.items():
            if not attr_value:
                continue
            if tag_lower == "a" and attr_name == "href":
                continue
            extracted_urls = extract_pdf_urls(attr_value)
            if extracted_urls:
                text = attrs_map.get("title") or attrs_map.get("aria-label") or attrs_map.get("alt") or ""
                for extracted in extracted_urls:
                    self.links.append((extracted, text))

        if tag_lower == "meta" and attrs_map.get("content") and "refresh" in attrs_map.get("http-equiv", "").lower():
            match = re.search(r"url\s*=\s*(.+)$", attrs_map["content"], flags=re.I)
            if match:
                self.links.append((match.group(1).strip(), ""))

        if tag_lower in {"script", "style"}:
            self._script_tag = tag_lower
            self._script_buffer = []

    def handle_data(self, data: str) -> None:
        if self._anchor_href is not None:
            self._anchor_text.append(data)
        if self._script_tag is not None:
            self._script_buffer.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag_lower = tag.lower()
        if tag_lower == "a" and self._anchor_href is not None:
            text = clean_whitespace("".join(self._anchor_text))
            if not text:
                text = clean_whitespace(
                    self._anchor_attrs.get("title")
                    or self._anchor_attrs.get("aria-label")
                    or self._anchor_attrs.get("data-title")
                    or self._anchor_attrs.get("download")
                )
            self.links.append((self._anchor_href, text))
            self._anchor_href = None
            self._anchor_text = []
            self._anchor_attrs = {}

        if tag_lower == self._script_tag and self._script_buffer:
            script_text = "".join(self._script_buffer)
            for extracted in extract_pdf_urls(script_text):
                self.links.append((extracted, "script"))
            self._script_tag = None
            self._script_buffer = []


def extract_pdf_urls(value: str) -> list[str]:
    if not value:
        return []
    candidates = re.findall(
        r"""(?i)(?:https?://|//|/|\.{1,2}/)[^"'<>]+?\.pdf(?:\?[^"'<>]*)?""",
        value,
    )
    cleaned: list[str] = []
    for candidate in candidates:
        candidate = candidate.replace("\\/", "/").strip().rstrip(".,;)")
        if candidate:
            cleaned.append(candidate)
    return cleaned


def clean_whitespace(value: str | None) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_company_numbering(value: str) -> str:
    return re.sub(r"^\d+[.]?\s*", "", clean_whitespace(value)).strip()


def ascii_fold(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return normalized.encode("ascii", "ignore").decode("ascii")


def normalize_company_name(value: str) -> str:
    value = ascii_fold(strip_company_numbering(value)).lower()
    value = re.sub(r"[*]+", "", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def safe_slug(value: str) -> str:
    value = ascii_fold(value).lower()
    value = re.sub(r"^\s*pt\.?\s+", "", value)
    value = re.sub(r"\b(persero|tbk)\b\.?", "", value)
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return re.sub(r"_+", "_", value).strip("_") or "company"


def normalize_source_url(value: str | None) -> str:
    url = clean_whitespace(value)
    if not url:
        return ""
    if url.lower() in INVALID_SOURCE_VALUES:
        return ""
    if not re.match(r"^[a-z]+://", url, flags=re.I):
        return ""
    return url


def detect_source_kind(source_url: str) -> str:
    lower = source_url.lower().strip()
    if not lower:
        return "invalid"
    if lower.endswith(".pdf") or ".pdf?" in lower:
        return "pdf"
    return "html"


def canonicalize_document_url(url: str) -> str:
    parsed = urlparse(url)
    return parsed._replace(fragment="").geturl().strip()


def summarize_exception(exc: Exception | str, limit: int = 240) -> str:
    text = clean_whitespace(str(exc).splitlines()[0] if str(exc) else exc if isinstance(exc, str) else repr(exc))
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def parse_target_period(value: str | None) -> tuple[int, int]:
    if not value:
        today = date.today()
        year = today.year if today.month > 1 else today.year - 1
        month = today.month - 1 if today.month > 1 else 12
        return year, month
    match = re.fullmatch(r"(\d{4})-(\d{2})", value.strip())
    if not match:
        raise ValueError("period must be YYYY-MM")
    return int(match.group(1)), int(match.group(2))


def parse_period_label(*candidates: str | None) -> tuple[int | None, int | None, str | None]:
    text = clean_whitespace(" ".join(ascii_fold(unquote(c or "")).lower() for c in candidates if c))
    if not text:
        return None, None, None

    month_map = {
        "januari": 1,
        "februari": 2,
        "maret": 3,
        "april": 4,
        "mei": 5,
        "juni": 6,
        "juli": 7,
        "agustus": 8,
        "september": 9,
        "oktober": 10,
        "november": 11,
        "desember": 12,
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12,
    }

    match = re.search(r"\b(20\d{2})[-_/](0?[1-9]|1[0-2])\b", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return year, month, f"{month:02d}-{year}"

    match = re.search(r"\b(0?[1-9]|1[0-2])[-_/](20\d{2})\b", text)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        return year, month, f"{month:02d}-{year}"

    match = re.search(r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return year, month, f"{month:02d}-{year}"

    year_match = re.search(r"(20\d{2})", text)
    year = int(year_match.group(1)) if year_match else None

    for token, month in month_map.items():
        if re.search(rf"\b{re.escape(token)}\b", text):
            label = f"{month:02d}-{year}" if year else None
            return year, month, label

    for token, month in month_map.items():
        if token in text:
            label = f"{month:02d}-{year}" if year else None
            return year, month, label

    return year, None, str(year) if year else None


def load_primary_registry(workbook_path: Path) -> list[tuple[tuple[str, str], str, str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        current_sector: str | None = None
        rows: list[tuple[tuple[str, str], str, str, str]] = []
        for row in sheet.iter_rows(values_only=True):
            first = row[0] if len(row) > 0 else None
            second = row[1] if len(row) > 1 else None
            third = row[2] if len(row) > 2 else None
            if isinstance(first, str) and first.strip() in SECTOR_FOLDER:
                current_sector = first.strip()
                continue
            if isinstance(first, (int, float)) and current_sector and second:
                company_name = clean_whitespace(str(second))
                key = (current_sector, normalize_company_name(company_name))
                rows.append((key, current_sector, company_name, normalize_source_url(str(third) if third is not None else "")))
        return rows
    finally:
        workbook.close()


def load_fallback_registry(workbook_path: Path) -> dict[tuple[str, str], str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        fallback: dict[tuple[str, str], str] = {}
        for row in sheet.iter_rows(min_row=3, values_only=True):
            for sector_label, company_col, url_col in FALLBACK_SHEET_GROUPS:
                company_cell = row[company_col] if len(row) > company_col else None
                url_cell = row[url_col] if len(row) > url_col else None
                if not company_cell:
                    continue
                company_name = strip_company_numbering(str(company_cell))
                source_url = normalize_source_url(str(url_cell) if url_cell is not None else "")
                key = (sector_label, normalize_company_name(company_name))
                if source_url and key not in fallback:
                    fallback[key] = source_url
        return fallback
    finally:
        workbook.close()


def load_company_registry(primary_workbook: Path, fallback_workbook: Path) -> list[CompanyRecord]:
    primary_rows = load_primary_registry(primary_workbook)
    fallback_urls = load_fallback_registry(fallback_workbook)

    records: list[CompanyRecord] = []
    for index, (key, sector_label, company_name, primary_url) in enumerate(primary_rows, start=1):
        fallback_url = fallback_urls.get(key, "")
        primary_source = None
        fallback_source = None
        if primary_url:
            primary_source = SourceEndpoint("primary", primary_url, detect_source_kind(primary_url))
        if fallback_url:
            fallback_source = SourceEndpoint("fallback", fallback_url, detect_source_kind(fallback_url))

        records.append(
            CompanyRecord(
                company_id=f"{SECTOR_FOLDER[sector_label]}-{index:03d}",
                company_name=company_name,
                company_type=SECTOR_FOLDER[sector_label],
                sector_label=sector_label,
                primary_source=primary_source,
                fallback_source=fallback_source,
            )
        )
    return records


def extract_candidate_links(
    html: str,
    company: CompanyRecord,
    source: SourceEndpoint,
    *,
    discovery_mode: str,
) -> list[DiscoveredDocument]:
    results: list[DiscoveredDocument] = []
    seen: set[str] = set()

    parser = LinkExtractor()
    parser.feed(html)
    parser.close()

    def add_candidate(raw_href: str, text: str | None = None) -> None:
        href = clean_whitespace(raw_href)
        if not href:
            return
        if href.lower().startswith(("javascript:", "mailto:", "tel:")):
            return
        resolved = canonicalize_document_url(urljoin(source.url, href))
        url_lower = resolved.lower()
        text_lower = clean_whitespace(text).lower()
        if url_lower in seen:
            return
        if ".pdf" not in url_lower and not any(token in text_lower for token in PDF_HINTS):
            return
        candidate_text = f"{text_lower} {url_lower}"
        if any(token in candidate_text for token in EXCLUDE_HINTS) and not any(token in candidate_text for token in MONTHLY_HINTS):
            return
        year, month, label = parse_period_label(resolved, text)
        results.append(
            DiscoveredDocument(
                company_id=company.company_id,
                company_name=company.company_name,
                company_type=company.company_type,
                source_url=source.url,
                source_label=source.label,
                source_kind=source.kind,
                discovery_mode=discovery_mode,
                document_url=resolved,
                anchor_text=text or None,
                report_year=year,
                report_month=month,
                report_period_label=label,
            )
        )
        seen.add(url_lower)

    for raw_href, text in parser.links:
        add_candidate(raw_href, text)

    for raw_url in extract_pdf_urls(html):
        add_candidate(raw_url, "raw_pdf_url")

    return results


def company_specific_candidates(
    company: CompanyRecord,
    source: SourceEndpoint,
    target_year: int,
    target_month: int,
) -> list[DiscoveredDocument]:
    host = urlparse(source.url).netloc.lower().removeprefix("www.")
    if "marein-re.com" not in host:
        return []

    month_name = MONTH_ID[target_month]
    base_url = "https://marein-re.com/public/uploads/finance_report/file"
    filenames = [
        f"LK_Publikasi_Marein_{month_name}_{target_year}.pdf",
        f"LK_Publikasi_Marein_{month_name.lower()}_{target_year}.pdf",
        f"LK_Publikasi_Marein_{month_name}{target_year}.pdf",
        f"LK_Publikasi_Marein_{target_year}_{month_name}.pdf",
    ]
    return [
        DiscoveredDocument(
            company_id=company.company_id,
            company_name=company.company_name,
            company_type=company.company_type,
            source_url=source.url,
            source_label=source.label,
            source_kind="pdf",
            discovery_mode="rule",
            document_url=f"{base_url}/{filename}",
            anchor_text=f"marein fallback {month_name} {target_year}",
            report_year=target_year,
            report_month=target_month,
            report_period_label=f"{target_month:02d}-{target_year}",
            notes="company-specific fallback",
        )
        for filename in filenames
    ]


def document_has_explicit_wrong_period(doc: DiscoveredDocument, target_year: int, target_month: int) -> bool:
    if doc.report_year is not None and doc.report_year != target_year:
        return True
    if doc.report_month is not None and doc.report_month != target_month:
        return True
    return False


def filter_target_documents(
    docs: list[DiscoveredDocument],
    target_year: int,
    target_month: int,
) -> list[DiscoveredDocument]:
    exact_docs = [doc for doc in docs if doc.report_year == target_year and doc.report_month == target_month]
    if exact_docs:
        return exact_docs

    compatible_docs = [doc for doc in docs if not document_has_explicit_wrong_period(doc, target_year, target_month)]
    if compatible_docs:
        return compatible_docs

    return []


def rank_document(doc: DiscoveredDocument, target_year: int, target_month: int) -> tuple[int, int, int, int, str]:
    text = ascii_fold(f"{doc.anchor_text or ''} {doc.document_url} {doc.notes or ''}").lower()

    if doc.report_year == target_year and doc.report_month == target_month:
        period_score = 0
    elif doc.report_year == target_year and doc.report_month is None:
        period_score = 1
    elif doc.report_year is None and doc.report_month is None:
        period_score = 2
    elif document_has_explicit_wrong_period(doc, target_year, target_month):
        period_score = 8
    else:
        period_score = 4

    kind_score = 0 if ".pdf" in doc.document_url.lower() or doc.source_kind == "pdf" else 1

    topic_score = 2
    if "syariah" in text or "sharia" in text:
        topic_score = 5
    elif "konvens" in text or "conventional" in text:
        topic_score = 0
    elif any(token in text for token in ("laporan keuangan", "financial report", "laporan bulanan", "monthly report")):
        topic_score = 0
    elif any(token in text for token in PDF_HINTS):
        topic_score = 1
    if any(token in text for token in EXCLUDE_HINTS) and not any(token in text for token in MONTHLY_HINTS):
        topic_score += 3

    source_score_map = {
        ("primary", "http"): 0,
        ("fallback", "http"): 1,
        ("primary", "browser"): 2,
        ("fallback", "browser"): 3,
        ("primary", "rule"): 4,
        ("fallback", "rule"): 5,
    }
    source_score = source_score_map.get((doc.source_label, doc.discovery_mode), 6)
    return (period_score, kind_score, topic_score, source_score, doc.document_url)


def looks_like_financial_report(doc: DiscoveredDocument) -> bool:
    text = ascii_fold(f"{doc.anchor_text or ''} {doc.document_url} {doc.notes or ''}").lower()
    positive = (
        "laporan keuangan",
        "financial report",
        "monthly report",
        "laporan bulanan",
        "financial statement",
        "newspaper",
        "publikasi",
        "report",
        "lapkeu",
        "bulanan",
    )
    negative = (
        "riplay",
        "kendaraan",
        "vehicle",
        "motor",
        "travel",
        "perjalanan",
        "kebakaran",
        "gempa",
        "home shield",
        "claim",
        "claims",
        "produk",
        "product",
    )
    if any(token in text for token in positive):
        return True
    if any(token in text for token in negative):
        return False
    return False


def is_pdf_payload(content_type: str, content_disposition: str, content: bytes) -> bool:
    lowered_content_type = clean_whitespace(content_type).lower()
    lowered_disposition = clean_whitespace(content_disposition).lower()
    return content.startswith(b"%PDF") or "pdf" in lowered_content_type or "pdf" in lowered_disposition


def is_pdf_response(response: Any) -> bool:
    headers = getattr(response, "headers", {}) or {}
    content = getattr(response, "content", b"") or b""
    return is_pdf_payload(headers.get("content-type", ""), headers.get("content-disposition", ""), content)


def ensure_httpx_available() -> Any:
    if httpx is None:
        raise MissingDependencyError("httpx is not installed. Install it before running downloads.")
    return httpx


def is_ssl_verification_error(exc: Exception) -> bool:
    return "CERTIFICATE_VERIFY_FAILED" in str(exc) or "certificate verify failed" in str(exc).lower()


def insecure_retry_response(client: Any, url: str, headers: dict[str, str]) -> Any:
    module = ensure_httpx_available()
    base_headers = dict(getattr(client, "headers", {}) or {})
    base_headers.update(headers)
    timeout = getattr(client, "timeout", None)
    with module.Client(
        headers=base_headers,
        follow_redirects=True,
        timeout=timeout,
        verify=False,
    ) as insecure_client:
        response = insecure_client.get(url)
        status_code = getattr(response, "status_code", 0)
        if status_code in BLOCKED_HTTP_STATUS_CODES:
            raise SourceBlockedError(f"HTTP {status_code} for {url}")
        response.raise_for_status()
        return response


def create_session(user_agent: str) -> Any:
    module = ensure_httpx_available()
    return module.Client(
        headers={
            "User-Agent": user_agent,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        },
        follow_redirects=True,
        timeout=module.Timeout(timeout=90.0, connect=20.0),
    )


def fetch_response(
    client: Any,
    url: str,
    *,
    retries: int,
    accept: str,
    referer: str | None = None,
) -> Any:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            headers = {"Accept": accept}
            if referer:
                headers["Referer"] = referer
            response = client.get(url, headers=headers)
            status_code = getattr(response, "status_code", 0)
            if status_code in BLOCKED_HTTP_STATUS_CODES:
                raise SourceBlockedError(f"HTTP {status_code} for {url}")
            response.raise_for_status()
            return response
        except SourceBlockedError:
            raise
        except Exception as exc:  # noqa: BLE001
            if is_ssl_verification_error(exc):
                try:
                    return insecure_retry_response(client, url, headers)
                except Exception as insecure_exc:  # noqa: BLE001
                    last_error = insecure_exc
                else:
                    return  # pragma: no cover - kept for flow clarity.
            else:
                last_error = exc
            if attempt < retries:
                time.sleep(min(2.0 * attempt, 5.0))
    assert last_error is not None
    raise last_error


def download_pdf(client: Any, url: str, output_path: Path, retries: int, referer: str | None = None) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".part")
    last_error: Exception | None = None

    for attempt in range(1, retries + 1):
        try:
            response = fetch_response(
                client,
                url,
                retries=1,
                accept="application/pdf,*/*;q=0.8",
                referer=referer,
            )
            if not is_pdf_response(response):
                raise ValueError(f"source did not return a PDF: {url}")
            content = response.content
            if len(content) < 1024:
                raise ValueError(f"received unusually small PDF payload: {url}")
            tmp_path.write_bytes(content)
            tmp_path.replace(output_path)
            return
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            tmp_path.unlink(missing_ok=True)
            if attempt < retries:
                time.sleep(min(2.0 * attempt, 5.0))

    assert last_error is not None
    raise last_error


def discover_documents_http(
    client: Any,
    company: CompanyRecord,
    source: SourceEndpoint,
    target_year: int,
    target_month: int,
    *,
    retries: int,
) -> list[DiscoveredDocument]:
    if source.kind == "invalid":
        return []
    if source.kind == "pdf":
        year, month, label = parse_period_label(source.url, company.company_name)
        return [
            DiscoveredDocument(
                company_id=company.company_id,
                company_name=company.company_name,
                company_type=company.company_type,
                source_url=source.url,
                source_label=source.label,
                source_kind="pdf",
                discovery_mode="http",
                document_url=source.url,
                anchor_text=None,
                report_year=year or target_year,
                report_month=month or target_month,
                report_period_label=label or f"{target_month:02d}-{target_year}",
            )
        ]

    response = fetch_response(
        client,
        source.url,
        retries=retries,
        accept="text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    )
    discovered = extract_candidate_links(response.text, company, source, discovery_mode="http")
    discovered.extend(company_specific_candidates(company, source, target_year, target_month))
    return expand_html_candidates_http(client, company, dedupe_documents(discovered), retries=retries, max_depth=2)


def dedupe_documents(docs: list[DiscoveredDocument]) -> list[DiscoveredDocument]:
    deduped: list[DiscoveredDocument] = []
    seen_urls: set[str] = set()
    for doc in docs:
        key = canonicalize_document_url(doc.document_url).lower()
        if key in seen_urls:
            continue
        seen_urls.add(key)
        deduped.append(doc)
    return deduped


def expand_html_candidates_http(
    client: Any,
    company: CompanyRecord,
    docs: list[DiscoveredDocument],
    *,
    retries: int,
    max_depth: int = 1,
) -> list[DiscoveredDocument]:
    expanded: list[DiscoveredDocument] = []
    queue: list[tuple[DiscoveredDocument, int]] = []
    visited_html: set[str] = set()

    for doc in docs:
        if doc.source_kind == "pdf" or ".pdf" in doc.document_url.lower():
            expanded.append(doc)
        else:
            queue.append((doc, 0))

    while queue:
        doc, depth = queue.pop(0)
        html_url = canonicalize_document_url(doc.document_url)
        if html_url.lower() in visited_html or depth >= max_depth:
            continue
        visited_html.add(html_url.lower())

        nested_source = SourceEndpoint(doc.source_label, html_url, "html")
        try:
            response = fetch_response(
                client,
                html_url,
                retries=retries,
                accept="text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                referer=doc.source_url,
            )
        except Exception:  # noqa: BLE001
            continue

        nested_docs = extract_candidate_links(response.text, company, nested_source, discovery_mode=doc.discovery_mode)
        for nested_doc in nested_docs:
            if nested_doc.source_kind == "pdf" or ".pdf" in nested_doc.document_url.lower():
                expanded.append(nested_doc)
            else:
                queue.append((nested_doc, depth + 1))

    return dedupe_documents(expanded)


class BrowserSession:
    def __init__(self, user_agent: str, timeout_ms: int) -> None:
        self.user_agent = user_agent
        self.timeout_ms = timeout_ms
        self._playwright = None
        self._browser = None
        self._context = None

    def __enter__(self) -> "BrowserSession":
        if sync_playwright is None:
            raise MissingDependencyError(
                "playwright is not installed. Install it and run `playwright install chromium`."
            )
        self._playwright = sync_playwright().start()
        self._browser = self._playwright.chromium.launch(headless=True)
        self._context = self._browser.new_context(user_agent=self.user_agent, accept_downloads=True)
        self._context.set_default_timeout(self.timeout_ms)
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        with suppress(Exception):
            if self._context is not None:
                self._context.close()
        with suppress(Exception):
            if self._browser is not None:
                self._browser.close()
        with suppress(Exception):
            if self._playwright is not None:
                self._playwright.stop()

    @property
    def ready(self) -> bool:
        return self._context is not None

    def discover_documents(
        self,
        company: CompanyRecord,
        source: SourceEndpoint,
        target_year: int,
        target_month: int,
    ) -> list[DiscoveredDocument]:
        if not self.ready:
            raise RuntimeError("browser session is not open")
        if source.kind == "pdf":
            year, month, label = parse_period_label(source.url, company.company_name)
            return [
                DiscoveredDocument(
                    company_id=company.company_id,
                    company_name=company.company_name,
                    company_type=company.company_type,
                    source_url=source.url,
                    source_label=source.label,
                    source_kind="pdf",
                    discovery_mode="browser",
                    document_url=source.url,
                    anchor_text=None,
                    report_year=year or target_year,
                    report_month=month or target_month,
                    report_period_label=label or f"{target_month:02d}-{target_year}",
                )
            ]

        page = self._context.new_page()
        downloads: list[tuple[str, str]] = []

        def remember_download(download: Any) -> None:
            with suppress(Exception):
                downloads.append((clean_whitespace(download.url), clean_whitespace(download.suggested_filename)))

        page.on("download", remember_download)
        try:
            page.goto(source.url, wait_until="domcontentloaded")
            with suppress(PlaywrightTimeoutError):
                page.wait_for_load_state("networkidle", timeout=min(self.timeout_ms, 8_000))

            html = page.content()
            docs = extract_candidate_links(html, company, source, discovery_mode="browser")

            link_rows = page.eval_on_selector_all(
                "a[href], iframe[src], embed[src], object[data], source[src]",
                """
                elements => elements.map((element) => {
                    const href = element.getAttribute('href')
                        || element.getAttribute('src')
                        || element.getAttribute('data')
                        || '';
                    const text = (element.innerText || element.textContent || element.getAttribute('title') || '').trim();
                    const container = element.closest('tr, li, article, section, div');
                    const context = ((container && (container.innerText || container.textContent)) || '').trim();
                    return { href, text, context };
                })
                """,
            )
            for row in link_rows:
                docs.extend(
                    extract_candidate_links(
                        f"<a href=\"{row.get('href', '')}\">{row.get('text', '')} {row.get('context', '')}</a>",
                        company,
                        source,
                        discovery_mode="browser",
                    )
                )

            if not docs:
                self._click_probable_downloads(page, company, source, downloads)

            for download_url, filename in downloads:
                year, month, label = parse_period_label(download_url, filename)
                candidate_url = download_url or page.url
                docs.append(
                    DiscoveredDocument(
                        company_id=company.company_id,
                        company_name=company.company_name,
                        company_type=company.company_type,
                        source_url=source.url,
                        source_label=source.label,
                        source_kind="pdf",
                        discovery_mode="browser",
                        document_url=canonicalize_document_url(candidate_url),
                        anchor_text=filename or "browser_download",
                        report_year=year,
                        report_month=month,
                        report_period_label=label,
                        notes="captured browser download event",
                    )
                )

            docs.extend(company_specific_candidates(company, source, target_year, target_month))
            return dedupe_documents(docs)
        finally:
            with suppress(Exception):
                page.close()

    def _click_probable_downloads(
        self,
        page: Any,
        company: CompanyRecord,
        source: SourceEndpoint,
        downloads: list[tuple[str, str]],
    ) -> None:
        selectors = [
            "a",
            "button",
            "[role='button']",
        ]
        texts = (
            "laporan",
            "financial",
            "report",
            "publikasi",
            "monthly",
            "download",
            "unduh",
        )
        for selector in selectors:
            locator = page.locator(selector)
            count = min(locator.count(), 12)
            for index in range(count):
                item = locator.nth(index)
                with suppress(Exception):
                    text = clean_whitespace(item.inner_text()).lower()
                    if not text or not any(token in text for token in texts):
                        continue
                    with page.expect_download(timeout=2_000) as download_info:
                        item.click()
                    download = download_info.value
                    downloads.append((clean_whitespace(download.url), clean_whitespace(download.suggested_filename)))
                    LOGGER.debug(
                        "captured browser download click for %s via %s/%s",
                        company.company_name,
                        source.label,
                        selector,
                    )

    def download_pdf(self, document: DiscoveredDocument, output_path: Path) -> None:
        if not self.ready:
            raise RuntimeError("browser session is not open")
        page = self._context.new_page()
        tmp_path = output_path.with_suffix(output_path.suffix + ".part")
        try:
            response = page.goto(document.document_url, wait_until="domcontentloaded")
            if response is None:
                raise ValueError(f"browser did not return a response for {document.document_url}")
            body = response.body()
            headers = response.headers or {}
            if not is_pdf_payload(headers.get("content-type", ""), headers.get("content-disposition", ""), body):
                raise ValueError(f"browser response did not return a PDF: {document.document_url}")
            if len(body) < 1024:
                raise ValueError(f"browser received unusually small PDF payload: {document.document_url}")
            output_path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path.write_bytes(body)
            tmp_path.replace(output_path)
        finally:
            tmp_path.unlink(missing_ok=True)
            with suppress(Exception):
                page.close()


def build_output_path(root: Path, year: int, month: int, company_type: str, company_name: str) -> Path:
    return root / "raw" / str(year) / f"{month:02d}" / company_type / f"{safe_slug(company_name)}_{year}_{month:02d}.pdf"


def setup_logging(output_root: Path, target_year: int, target_month: int, verbose: bool) -> Path:
    log_dir = output_root / "logs" / str(target_year) / f"{target_month:02d}"
    log_dir.mkdir(parents=True, exist_ok=True)
    run_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_log = log_dir / f"download_market_reports_{run_stamp}.log"

    LOGGER.handlers.clear()
    LOGGER.setLevel(logging.DEBUG if verbose else logging.INFO)
    LOGGER.propagate = False

    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    stream_handler.setLevel(logging.DEBUG if verbose else logging.INFO)
    file_handler = logging.FileHandler(run_log, encoding="utf-8")
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG if verbose else logging.INFO)

    LOGGER.addHandler(stream_handler)
    LOGGER.addHandler(file_handler)
    LOGGER.info("writing run log to %s", run_log)
    return run_log


def company_status_line(index: int, total: int, company: CompanyRecord, output_path: Path) -> None:
    LOGGER.info("[%d/%d] %s -> %s", index, total, company.company_name, output_path)


def discover_all_documents(
    client: Any,
    browser: BrowserSession | None,
    company: CompanyRecord,
    target_year: int,
    target_month: int,
    *,
    retries: int,
    browser_enabled: bool,
) -> tuple[list[DiscoveredDocument], list[str]]:
    documents: list[DiscoveredDocument] = []
    attempts: list[str] = []

    for source in company.ordered_sources():
        try:
            http_docs = discover_documents_http(client, company, source, target_year, target_month, retries=retries)
            if http_docs:
                documents.extend(http_docs)
            attempts.append(f"{source.label}:http:ok")
        except SourceBlockedError as exc:
            attempts.append(f"{source.label}:http:blocked:{summarize_exception(exc)}")
        except Exception as exc:  # noqa: BLE001
            attempts.append(f"{source.label}:http:error:{summarize_exception(exc)}")

    http_documents = dedupe_documents(documents)
    http_exact_documents = [
        doc
        for doc in http_documents
        if doc.report_year == target_year and doc.report_month == target_month and looks_like_financial_report(doc)
    ]
    if http_exact_documents:
        return filter_target_documents(http_documents, target_year, target_month), attempts

    if not browser_enabled:
        return filter_target_documents(http_documents, target_year, target_month), attempts

    if browser is None:
        attempts.append("browser:unavailable")
        return filter_target_documents(http_documents, target_year, target_month), attempts

    for source in company.ordered_sources():
        try:
            browser_docs = browser.discover_documents(company, source, target_year, target_month)
            if browser_docs:
                documents.extend(browser_docs)
            attempts.append(f"{source.label}:browser:ok")
        except Exception as exc:  # noqa: BLE001
            attempts.append(f"{source.label}:browser:error:{summarize_exception(exc)}")

    filtered_documents = filter_target_documents(dedupe_documents(documents), target_year, target_month)
    return filtered_documents, attempts


def write_failure_log(path: Path, rows: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if rows:
        path.write_text("\n".join(rows) + "\n", encoding="utf-8")
    else:
        path.write_text("", encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Download monthly insurer reports into raw/YYYY/MM/... folders.")
    parser.add_argument("--workbook-primary", type=Path, default=Path("asset_source/URL Perusahaan Asuransi.xlsx"))
    parser.add_argument("--workbook-fallback", type=Path, default=Path("asset_source/List Link Lapkeu.xlsx"))
    parser.add_argument("--output-root", type=Path, default=Path("data"))
    parser.add_argument("--period", type=str, default="2026-03")
    parser.add_argument("--sector", type=str, default=None)
    parser.add_argument("--company", type=str, default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--retries", type=int, default=5)
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--delay", type=float, default=0.25)
    parser.add_argument("--user-agent", type=str, default="Mozilla/5.0")
    parser.add_argument("--browser-timeout", type=int, default=30000)
    parser.add_argument("--disable-browser-fallback", action="store_true")
    return parser.parse_args(argv)


def run(argv: list[str]) -> int:
    args = parse_args(argv)

    if not args.workbook_primary.exists():
        print(f"ERROR workbook not found: {args.workbook_primary}", file=sys.stderr)
        return 2
    if not args.workbook_fallback.exists():
        print(f"ERROR workbook not found: {args.workbook_fallback}", file=sys.stderr)
        return 2

    target_year, target_month = parse_target_period(args.period)
    run_log = setup_logging(args.output_root, target_year, target_month, args.verbose)

    records = load_company_registry(args.workbook_primary, args.workbook_fallback)
    if len(records) != EXPECTED_COMPANY_COUNT:
        LOGGER.error("registry count mismatch: expected=%d actual=%d", EXPECTED_COMPANY_COUNT, len(records))
        return 2

    if args.sector:
        sector_query = args.sector.lower().strip()
        records = [
            record
            for record in records
            if record.company_type == sector_query
            or record.sector_label.lower() == sector_query
            or SECTOR_FOLDER.get(args.sector, "") == record.company_type
        ]
    if args.company:
        company_query = args.company.lower().strip()
        records = [record for record in records if company_query in record.company_name.lower()]
    if args.limit is not None:
        records = records[: args.limit]

    LOGGER.info("target period %04d-%02d, companies=%d, dry_run=%s", target_year, target_month, len(records), args.dry_run)
    LOGGER.info("primary workbook %s", args.workbook_primary)
    LOGGER.info("fallback workbook %s", args.workbook_fallback)

    if args.dry_run:
        for index, company in enumerate(records, start=1):
            output_path = build_output_path(args.output_root, target_year, target_month, company.company_type, company.company_name)
            company_status_line(index, len(records), company, output_path)
            LOGGER.info("  dry-run")
        LOGGER.info("done: total=%d saved=0 existing=0 failed=0", len(records))
        return 0

    if httpx is None:
        LOGGER.error("httpx is not installed. Install it before running downloads. log=%s", run_log)
        return 2

    browser_enabled = not args.disable_browser_fallback
    browser_missing: str | None = None
    failures_reported: list[str] = []
    saved = 0
    existing = 0
    failed = 0

    error_log = args.output_root / "raw" / str(target_year) / f"{target_month:02d}" / "error.log"

    browser_context = None
    if browser_enabled:
        try:
            browser_context = BrowserSession(args.user_agent, args.browser_timeout)
            browser_context.__enter__()
        except MissingDependencyError as exc:
            browser_missing = summarize_exception(exc)
            browser_context = None
            LOGGER.warning("browser fallback disabled: %s", browser_missing)
        except Exception as exc:  # noqa: BLE001
            browser_missing = summarize_exception(exc)
            browser_context = None
            LOGGER.warning("browser fallback disabled: %s", browser_missing)

    try:
        with create_session(args.user_agent) as session:
            total = len(records)
            for index, company in enumerate(records, start=1):
                output_path = build_output_path(args.output_root, target_year, target_month, company.company_type, company.company_name)
                company_status_line(index, total, company, output_path)

                if output_path.exists() and not args.overwrite:
                    LOGGER.info("  existing pdf")
                    existing += 1
                    continue

                try:
                    documents, attempts = discover_all_documents(
                        session,
                        browser_context,
                        company,
                        target_year,
                        target_month,
                        retries=args.retries,
                        browser_enabled=browser_enabled,
                    )
                    documents = sorted(documents, key=lambda doc: rank_document(doc, target_year, target_month))

                    if not documents:
                        reasons = "; ".join(attempts) if attempts else "no candidate discovery attempts"
                        if browser_enabled and browser_missing:
                            reasons = f"{reasons}; browser unavailable: {browser_missing}"
                        raise RuntimeError(f"no March 2026 candidate found ({reasons})")

                    saved_this_company = False
                    candidate_errors: list[str] = []
                    for document in documents:
                        LOGGER.info(
                            "  trying candidate [%s/%s] %s",
                            document.source_label,
                            document.discovery_mode,
                            document.document_url,
                        )
                        try:
                            download_pdf(session, document.document_url, output_path, retries=args.retries, referer=document.source_url)
                            LOGGER.info("  saved")
                            saved += 1
                            saved_this_company = True
                            break
                        except SourceBlockedError as exc:
                            candidate_errors.append(f"{document.document_url} -> {summarize_exception(exc)}")
                            if browser_context is not None:
                                try:
                                    browser_context.download_pdf(document, output_path)
                                    LOGGER.info("  saved")
                                    saved += 1
                                    saved_this_company = True
                                    break
                                except Exception as browser_exc:  # noqa: BLE001
                                    candidate_errors.append(
                                        f"{document.document_url} -> browser {summarize_exception(browser_exc)}"
                                    )
                        except Exception as exc:  # noqa: BLE001
                            candidate_errors.append(f"{document.document_url} -> {summarize_exception(exc)}")

                    if not saved_this_company:
                        raise RuntimeError("; ".join(candidate_errors) or "all candidates failed")

                    if args.delay:
                        time.sleep(args.delay)
                except Exception as exc:  # noqa: BLE001
                    failed += 1
                    summary = summarize_exception(exc, limit=400)
                    LOGGER.error("  error: %s", summary)
                    primary_url = company.primary_source.url if company.primary_source else "-"
                    fallback_url = company.fallback_source.url if company.fallback_source else "-"
                    failures_reported.append(
                        "\t".join(
                            [
                                company.company_name,
                                company.sector_label,
                                primary_url,
                                fallback_url,
                                summary,
                            ]
                        )
                    )
    finally:
        if browser_context is not None:
            browser_context.__exit__(None, None, None)

    write_failure_log(error_log, failures_reported)
    if failures_reported:
        LOGGER.info("wrote failure log to %s", error_log)

    LOGGER.info("done: total=%d saved=%d existing=%d failed=%d", len(records), saved, existing, failed)
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(run(sys.argv[1:]))
