"""Fill the OJK standardized general insurance template with values from PDF.

Template value columns:
  Assets      → G (current), H (prior)     | labels in column B, rows 12-33
  Liabilities → O (current), P (prior)     | labels in column J, rows 12-67
  Income      → X (current), Y (prior)     | labels in column S, rows 12-75
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from copy import copy

TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "laporan_keuangan_template_asuransi_umum.xlsx"

_VAL_COLS = {
    "assets":      ("G",  "H"),
    "liabilities": ("O",  "P"),
    "income":      ("X",  "Y"),
}
_LABEL_COLS = {
    "assets":      "B",
    "liabilities": "J",
    "income":      "S",
}
_DATA_ROWS = range(12, 76)

# Synonym replacements applied BEFORE label matching (both sides normalised)
_SYNONYMS = {
    "utang piutang": "tagihan",
    "piutang": "tagihan",
    "komisi": "komisi",
    "hutang": "utang",
    "kas bank": "kas dan bank",
    "koasuransi": "koasuransi",
    "klaim": "klaim",
    "pajak": "pajak",
    "biaya": "biaya",
}


# ── Normalisation ────────────────────────────────────────────────────────────

def _norm(text: str) -> str:
    """Lowercase, strip item-number prefixes + parentheticals, apply synonyms."""
    if not text:
        return ""
    t = str(text).strip()
    # Strip leading roman numerals: I, II, III, IV, V, VI ...
    t = re.sub(r"^(IV|IX|V?I{1,3}|XI{0,2})\s*\.?\s+", "", t, flags=re.IGNORECASE)
    # Strip leading arabic item numbers:  "1.", "21.", "33 "
    t = re.sub(r"^\d+\s*\.?\s+", "", t)
    # Strip leading sub-item letters: "a.", "b.", "c.", "A.", "B."
    t = re.sub(r"^[a-eA-E]\s*\.?\s+", "", t)
    # Strip ALL parentheticals (formula refs AND semantic ones)
    t = re.sub(r"\s*\([^)]*\)", "", t)
    # Strip asterisks (template uses *) as footnote markers)
    t = t.replace("*", "")
    t = " ".join(t.split()).lower()
    for src, dst in _SYNONYMS.items():
        t = t.replace(src, dst)
    return t


# ── Value parsing ────────────────────────────────────────────────────────────

_NUM_RE = re.compile(r"^\(?-?[\d.,]+\)?%?$")


def _parse_val(text: str, thousands_sep: str = ".") -> float | None:
    """Parse Indonesian or US number format → float.

    Indonesian (thousands_sep="."):
      "1.234.567"     → 1234567.0   (dots = thousands)
      "1.234.567,89"  → 1234567.89  (comma = decimal)

    US (thousands_sep=","):
      "1,234,567"     → 1234567.0   (commas = thousands)
      "1,234,567.89"  → 1234567.89  (dot = decimal)

    Both:
      "(2.345)"       → -2345.0     (parentheses = negative)
      "93,3%"         → 93.3        (percentage; caller decides semantics)
      "-"             → None
    """
    if not text:
        return None
    t = str(text).strip()
    if t in ("-", ""):
        return None
    negative = t.startswith("(") and t.endswith(")")
    if negative:
        t = t[1:-1]
    t = t.rstrip("%").replace(" ", "")
    if not t:
        return None

    decimal_sep = "," if thousands_sep == "." else "."

    if decimal_sep in t:
        # has a decimal part
        parts = t.rsplit(decimal_sep, 1)
        try:
            val = float(parts[0].replace(thousands_sep, "") + "." + parts[1])
        except (ValueError, IndexError):
            return None
    else:
        try:
            val = float(t.replace(thousands_sep, ""))
        except ValueError:
            return None
    return -val if negative else val


def _split_two(s: str) -> tuple[str, str]:
    """Split a string that contains two space-separated numbers."""
    if not s:
        return "", ""
    parts = s.split()
    if len(parts) >= 2 and _NUM_RE.match(parts[0]) and _NUM_RE.match(parts[-1]):
        return parts[0], parts[-1]
    return s, ""


# ── Template index ───────────────────────────────────────────────────────────

def _build_index(ws) -> dict[str, dict[str, int]]:
    """Build {section → {normalized_label → row_number}} from the template sheet."""
    idx: dict[str, dict[str, int]] = {s: {} for s in _LABEL_COLS}
    for row in _DATA_ROWS:
        for sec, col in _LABEL_COLS.items():
            v = ws[f"{col}{row}"].value
            if v:
                n = _norm(str(v))
                if n and n not in idx[sec]:   # first occurrence wins
                    idx[sec][n] = row
    return idx


def _find_row(
    label: str,
    section: str,
    idx: dict[str, dict[str, int]],
    ovr: dict[str, dict[str, int]],
) -> int | None:
    n = _norm(label)
    if not n:
        return None
    # 1. Per-company override (keys already normalised)
    if n in ovr.get(section, {}):
        return ovr[section][n]
    # 2. Exact normalised match
    if n in idx[section]:
        return idx[section][n]
    # 3. Extracted label is contained in template label
    for t, row in idx[section].items():
        if n in t:
            return row
    # 4. Template label is contained in extracted label
    for t, row in idx[section].items():
        if t in n:
            return row
    return None


# ── Main fill function ───────────────────────────────────────────────────────

def fill_template(
    rows: list[dict],
    config: dict,
    period: str,
    out_path: Path,
) -> None:
    """Copy template, fill extracted numbers, save to out_path."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE_PATH, out_path)

    wb = load_workbook(out_path)
    ws = wb.active

    # ── Build label index + normalised overrides ──────────────────────────────
    idx = _build_index(ws)
    raw_ovr = config.get("label_map", {})
    ovr: dict[str, dict[str, int]] = {
        sec: {_norm(k): v for k, v in mapping.items()}
        for sec, mapping in raw_ovr.items()
    }

    # ── Zone config: which extracted columns hold label / current / prior ──────
    defaults = {
        "assets":      ("B", "C", "D"),
        "liabilities": ("F", "G", "H"),
        "income":      ("J", "K", "L"),
    }
    zcfg = config.get("zones", {})
    zones: dict[str, tuple[str, str, str] | None] = {}
    for sec, (dl, dc, dp) in defaults.items():
        zs = zcfg.get(sec)
        if zs is False or zs == "false":
            zones[sec] = None               # section absent in this company's PDF
        elif isinstance(zs, dict):
            zones[sec] = (zs.get("label_col", dl),
                          zs.get("val_current", dc),
                          zs.get("val_prior",   dp))
        else:
            zones[sec] = (dl, dc, dp)       # use defaults

    # ── Fill data rows ────────────────────────────────────────────────────────
    tsep = config.get("thousands_sep", ".")
    skip_labels_list = config.get("skip_labels", [])
    skip_labels_norm = {_norm(lbl) for lbl in skip_labels_list}

    unmatched: list[tuple[str, str]] = []
    for row in rows:
        for sec, zone in zones.items():
            if zone is None:
                continue
            lc, vc, vp = zone
            label = row.get(lc, "")
            if not label:
                continue

            # Skip rows that match skip_labels
            if _norm(label) in skip_labels_norm:
                continue

            raw_cur = row.get(vc, "")
            raw_pri = row.get(vp, "")

            # Both period values may have collapsed into the current column
            if raw_cur and not raw_pri:
                raw_cur, raw_pri = _split_two(raw_cur)

            val_cur = _parse_val(raw_cur, tsep)
            val_pri = _parse_val(raw_pri, tsep)
            if val_cur is None and val_pri is None:
                continue

            tmpl_row = _find_row(label, sec, idx, ovr)
            if tmpl_row is None:
                unmatched.append((sec, label))
                continue

            col_cur, col_pri = _VAL_COLS[sec]
            if val_cur is not None:
                ws[f"{col_cur}{tmpl_row}"] = val_cur
            if val_pri is not None:
                ws[f"{col_pri}{tmpl_row}"] = val_pri

    if unmatched:
        seen: set[tuple[str, str]] = set()
        cid = config.get("company_id", "?")
        print(f"  [WARN] {cid}: {len({k for k in unmatched})} unmatched label(s):")
        for sec, lbl in unmatched:
            if (sec, lbl) not in seen:
                print(f"    {sec}: {lbl!r}")
                seen.add((sec, lbl))

    wb.save(out_path)
