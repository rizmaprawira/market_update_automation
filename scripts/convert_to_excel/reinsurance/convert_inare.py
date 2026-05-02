"""Convert PT Indoperkasa Suksesjaya Reasuransi (INARE) monthly PDF to Excel.

Usage:
    conda run -n market_update python scripts/convert_to_excel/reinsurance/convert_inare.py

Input:  data/2026-03/raw_pdf/reasuransi/inare/inare_2026-03.pdf
Output: data/2026-03/raw_excel/reasuransi/inare/inare_2026-03.xlsx
"""
from __future__ import annotations

import subprocess
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROJECT_ROOT = Path(__file__).resolve().parents[3]
PDF_PATH = PROJECT_ROOT / "data/2026-03/raw_pdf/reasuransi/inare/inare_2026-03.pdf"
OUT_PATH = PROJECT_ROOT / "data/2026-03/raw_excel/reasuransi/inare/inare_2026-03.xlsx"

CONTENT_START_Y = 190.0
ROW_CLUSTER_TOL = 0.75

BODY_COLS = ("B", "C", "D", "F", "G", "H", "J", "K", "L", "N", "O", "P")
VALUE_COLS = {"C", "D", "G", "H", "K", "L", "O", "P"}

# A3 landscape (1190 × 842). Calibrated from pdftotext -bbox-layout.
# B < 230: labels; C 230-290: year-1 values; D 290-340: year-2 values;
# F 340-460: liab labels; G 460-530: liab year-1; H 530-590: liab year-2;
# J 590-700: income labels; K 700-790: income year-1; L 790-940: income year-2;
# N 940-990: ratio labels; O 990-1095: ratio year-1; P 1095+: ratio year-2.
X_THRESHOLDS = (230, 290, 320, 460, 530, 575, 700, 790, 870, 990, 1095)

BOLD_TOKENS = (
    "JUMLAH", "TOTAL", "LABA", "INVESTASI", "BUKAN INVESTASI",
    "KOMISARIS", "DIREKSI", "PEMILIK PERUSAHAAN", "ANAK PERUSAHAAN",
    "REASURADUR UTAMA", "KETERANGAN", "CATATAN", "SOLVABILITAS",
)
CENTER_TOKENS = (
    "KETERANGAN", "SOLVABILITAS", "RASIO", "INDIKATOR",
    "KOMISARIS DAN DIREKSI", "DEWAN KOMISARIS", "DIREKSI",
    "PEMILIK PERUSAHAAN", "ANAK PERUSAHAAN", "REASURADUR UTAMA",
    "CATATAN", "INFORMASI",
)


def parse_words(pdf_path: Path) -> list[dict]:
    xml = subprocess.check_output(["pdftotext", "-bbox-layout", str(pdf_path), "-"], text=True)
    root = ET.fromstring(xml)
    ns = {"x": "http://www.w3.org/1999/xhtml"}
    lines = []
    for line in root.findall(".//x:line", ns):
        words = []
        for w in line.findall("x:word", ns):
            text = (w.text or "").strip()
            if text:
                words.append({"x": float(w.attrib["xMin"]), "text": text})
        if words:
            words.sort(key=lambda ww: ww["x"])
            lines.append({"y": float(line.attrib["yMin"]), "words": words})
    lines.sort(key=lambda ll: (ll["y"], ll["words"][0]["x"]))
    return lines


def assign_col(x: float) -> str:
    for thresh, col in zip(X_THRESHOLDS, BODY_COLS[:-1]):
        if x < thresh:
            return col
    return BODY_COLS[-1]


def clean(text: str) -> str:
    return " ".join(text.split())


def row_flags(texts: list[str]) -> dict:
    joined = " | ".join(texts).upper()
    return {
        "bold": any(t in joined for t in BOLD_TOKENS)
                or joined.startswith(("I.", "II.", "III.", "IV.")),
        "center": any(t in joined for t in CENTER_TOKENS),
    }


def style(cell, *, font=None, fill=None, align=None, border=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(PDF_PATH)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    lines = parse_words(PDF_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "inare"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 55
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    col_widths = {
        "A": 4.0, "B": 36.0, "C": 15.0, "D": 15.0, "E": 2.0,
        "F": 34.0, "G": 15.0, "H": 15.0, "I": 2.0,
        "J": 40.0, "K": 15.0, "L": 15.0, "M": 2.0,
        "N": 44.0, "O": 15.0, "P": 15.0,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    f_title = Font(name="Times New Roman", size=13, bold=True)
    f_small = Font(name="Times New Roman", size=8)
    f_report = Font(name="Times New Roman", size=12, bold=True)
    f_period = Font(name="Times New Roman", size=9, bold=True)
    f_section = Font(name="Times New Roman", size=7, bold=True, color="FFFFFF")
    f_table = Font(name="Times New Roman", size=7, bold=True)
    f_year = Font(name="Times New Roman", size=7, bold=True)
    f_body = Font(name="Times New Roman", size=6.5)
    f_bold = Font(name="Times New Roman", size=6.5, bold=True)

    a_ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
    a_rgt = Alignment(horizontal="right", vertical="center", wrap_text=False)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_box = Border(left=medium, right=medium, top=medium, bottom=medium)

    fill_section = PatternFill("solid", fgColor="5B7B57")  # green
    fill_header = PatternFill("solid", fgColor="DDEAD7")
    fill_year = PatternFill("solid", fgColor="EEF5EA")
    fill_subtle = PatternFill("solid", fgColor="F2F2F2")

    for rng in ["A1:P1", "A2:P2", "A3:P3", "A5:P5", "A6:P6"]:
        ws.merge_cells(rng)
    ws["A1"] = "PT INDOPERKASA SUKSESJAYA REASURANSI"
    style(ws["A1"], font=f_title, align=a_ctr)
    ws.row_dimensions[1].height = 18
    ws["A2"] = "Kantor Pusat"
    style(ws["A2"], font=f_small, align=a_ctr)
    ws.row_dimensions[2].height = 13
    ws["A3"] = "Plaza Mutiara Lt.21, Jl. Dr. Ide Agung Anak Gde Agung, Kuningan Timur, Setiabudi, Jakarta Selatan 12950"
    style(ws["A3"], font=f_small, align=a_ctr)
    ws.row_dimensions[3].height = 13
    ws["A5"] = "LAPORAN BULANAN"
    style(ws["A5"], font=f_report, align=a_ctr)
    ws.row_dimensions[5].height = 18
    ws["A6"] = "Per 31 Maret 2026 dan 2025"
    style(ws["A6"], font=f_period, align=a_ctr)
    ws.row_dimensions[6].height = 15

    for rng in ["A7:D9", "J7:L9", "N7:P9"]:
        ws.merge_cells(rng)
    ws["A7"] = "LAPORAN POSISI KEUANGAN\n(dalam jutaan Rupiah)"
    ws["J7"] = "LAPORAN LABA (RUGI) KOMPREHENSIF\n(dalam jutaan Rupiah)"
    ws["N7"] = "INDIKATOR KESEHATAN KEUANGAN\n(dalam jutaan Rupiah)"
    for ref in ("A7", "J7", "N7"):
        style(ws[ref], font=f_section, fill=fill_section, align=a_ctr, border=b_box)
    ws.row_dimensions[7].height = 17
    ws.row_dimensions[8].height = 13
    ws.row_dimensions[9].height = 13

    for rng in ["A10:D10", "F10:H10", "J10:L10", "N10:P10"]:
        ws.merge_cells(rng)
    ws["A10"] = "ASET"
    ws["F10"] = "LIABILITAS DAN EKUITAS"
    ws["J10"] = "URAIAN"
    ws["N10"] = "URAIAN"
    for ref in ("A10", "F10", "J10", "N10"):
        style(ws[ref], font=f_table, fill=fill_header, align=a_ctr, border=b_thin)
    ws.row_dimensions[10].height = 18

    year_labels = {
        "C11": "2026", "D11": "2025",
        "G11": "2026", "H11": "2025",
        "K11": "2026", "L11": "2025",
        "O11": "2026", "P11": "2025",
    }
    for ref, text in year_labels.items():
        ws[ref] = text
        style(ws[ref], font=f_year, fill=fill_year, align=a_ctr, border=b_thin)
    ws.row_dimensions[11].height = 15

    content = [ln for ln in lines if ln["y"] >= CONTENT_START_Y]
    clusters: list[dict] = []
    for ln in content:
        if not clusters or abs(ln["y"] - clusters[-1]["min_y"]) > ROW_CLUSTER_TOL:
            clusters.append({"lines": [ln], "min_y": ln["y"], "max_y": ln["y"]})
        else:
            clusters[-1]["lines"].append(ln)
            clusters[-1]["min_y"] = min(clusters[-1]["min_y"], ln["y"])
            clusters[-1]["max_y"] = max(clusters[-1]["max_y"], ln["y"])

    start_row = 12
    for idx, grp in enumerate(clusters):
        row_num = start_row + idx
        gap = (clusters[idx + 1]["min_y"] - grp["min_y"]) if idx < len(clusters) - 1 else 8.0
        ws.row_dimensions[row_num].height = max(7.0, gap * 1.15)

        parts: dict[str, list[str]] = defaultdict(list)
        for ln in grp["lines"]:
            for w in ln["words"]:
                parts[assign_col(w["x"])].append(w["text"])

        texts = [clean(" ".join(v)) for v in parts.values()]
        flags = row_flags(texts)

        for col, words in parts.items():
            val = clean(" ".join(words))
            if val:
                ws[f"{col}{row_num}"] = val

        for col in BODY_COLS:
            cell = ws[f"{col}{row_num}"]
            if cell.value is None:
                continue
            is_val = col in VALUE_COLS
            style(
                cell,
                font=f_bold if flags["bold"] else f_body,
                align=a_rgt if is_val else (a_ctr if flags["center"] else a_lft),
                border=b_thin,
                fill=fill_subtle if flags["bold"] and not is_val else None,
            )

    last_row = start_row + len(clusters) - 1
    ws.print_area = f"A1:P{last_row}"
    ws.freeze_panes = "A12"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
