"""Convert PT Reasuransi Indonesia Utama (Persero) monthly PDF to Excel.

Usage:
    conda run -n market_update python scripts/convert_to_excel/reinsurance/convert_indonesiare.py

Input:  data/2026-03/raw_pdf/reasuransi/indonesiare/indonesiare_2026-03.pdf
Output: data/2026-03/raw_excel/reasuransi/indonesiare/indonesiare_2026-03.xlsx
"""
from __future__ import annotations

import subprocess
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROJECT_ROOT = Path(__file__).resolve().parents[3]
PDF_PATH = PROJECT_ROOT / "data/2026-03/raw_pdf/reasuransi/indonesiare/indonesiare_2026-03.pdf"
OUT_PATH = PROJECT_ROOT / "data/2026-03/raw_excel/reasuransi/indonesiare/indonesiare_2026-03.xlsx"

CONTENT_START_Y = 97.0
ROW_CLUSTER_TOL = 0.75

# Column B through P (skipping spacers E, I, M)
BODY_COLS = ("B", "C", "D", "F", "G", "H", "J", "K", "L", "N", "O", "P")
VALUE_COLS = {"C", "D", "G", "H", "K", "L", "O", "P"}

# x thresholds: 11 values map x-coordinate → one of 12 body columns
# Derived from pdftotext -bbox-layout word positions in the 2026-03 PDF
X_THRESHOLDS = (180, 215, 244, 380, 415, 440, 545, 575, 605, 705, 730)

BOLD_TOKENS = (
    "JUMLAH", "TOTAL", "LABA", "KOMISARIS", "DIREKSI",
    "PEMILIK PERUSAHAAN", "ANAK PERUSAHAAN", "REASURADUR UTAMA",
    "KETERANGAN", "INFORMASI LAIN", "CATATAN", "SOLVABILITAS",
    "INVESTASI", "BUKAN INVESTASI", "CADANGAN TEKNIS",
)
CENTER_TOKENS = (
    "KETERANGAN", "SOLVABILITAS", "PENCAPAIAN", "RASIO",
    "KOMISARIS DAN DIREKSI", "DEWAN KOMISARIS", "DIREKSI",
    "PEMILIK PERUSAHAAN", "ANAK PERUSAHAAN", "REASURADUR UTAMA",
    "INFORMASI LAIN", "CATATAN", "DANANTARA", "INDONESIARE",
)


def parse_words(pdf_path: Path) -> list[dict]:
    xml = subprocess.check_output(["pdftotext", "-bbox-layout", str(pdf_path), "-"], text=True)
    root = ET.fromstring(xml)
    ns = {"x": "http://www.w3.org/1999/xhtml"}
    lines = []
    for line in root.findall(".//x:line", ns):
        words = []
        for w in line.findall("x:word", ns):
            text = (w.text or "").strip()
            if text:
                words.append({"x": float(w.attrib["xMin"]), "text": text})
        if words:
            words.sort(key=lambda ww: ww["x"])
            lines.append({
                "y": float(line.attrib["yMin"]),
                "words": words,
            })
    lines.sort(key=lambda ll: (ll["y"], ll["words"][0]["x"]))
    return lines


def assign_col(x: float) -> str:
    for thresh, col in zip(X_THRESHOLDS, BODY_COLS[:-1]):
        if x < thresh:
            return col
    return BODY_COLS[-1]


def clean(text: str) -> str:
    return " ".join(text.split())


def row_flags(texts: list[str]) -> dict:
    joined = " | ".join(texts).upper()
    return {
        "bold": any(t in joined for t in BOLD_TOKENS)
                or joined.startswith(("I.", "II.", "III.", "IV.")),
        "center": any(t in joined for t in CENTER_TOKENS),
    }


def style(cell, *, font=None, fill=None, align=None, border=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(PDF_PATH)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    lines = parse_words(PDF_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "indonesiare"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 70
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    col_widths = {
        "A": 4.0, "B": 28.5, "C": 13.0, "D": 13.0, "E": 2.0,
        "F": 26.0, "G": 13.0, "H": 13.0, "I": 2.0,
        "J": 34.0, "K": 13.0, "L": 13.0, "M": 2.0,
        "N": 35.0, "O": 13.0, "P": 13.0,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Styles
    f_title = Font(name="Times New Roman", size=13, bold=True)
    f_small = Font(name="Times New Roman", size=8)
    f_report = Font(name="Times New Roman", size=12, bold=True)
    f_period = Font(name="Times New Roman", size=9, bold=True)
    f_section = Font(name="Times New Roman", size=7, bold=True, color="FFFFFF")
    f_table = Font(name="Times New Roman", size=7, bold=True)
    f_year = Font(name="Times New Roman", size=7, bold=True)
    f_body = Font(name="Times New Roman", size=6.5)
    f_bold = Font(name="Times New Roman", size=6.5, bold=True)

    a_ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
    a_rgt = Alignment(horizontal="right", vertical="center", wrap_text=False)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_box = Border(left=medium, right=medium, top=medium, bottom=medium)

    fill_black = PatternFill("solid", fgColor="000000")
    fill_light = PatternFill("solid", fgColor="F2F2F2")
    fill_grey = PatternFill("solid", fgColor="D9D9D9")

    # Header rows 1-6
    for rng in ["A1:P1", "A2:P2", "A3:P3", "A5:P5", "A6:P6"]:
        ws.merge_cells(rng)
    ws["A1"] = "PT REASURANSI INDONESIA UTAMA (PERSERO)"
    style(ws["A1"], font=f_title, align=a_ctr)
    ws.row_dimensions[1].height = 18
    ws["A2"] = "Telp : +62-21-3920101, 31934208 (hunting), P.O.Box 2635 Jkt 10026"
    style(ws["A2"], font=f_small, align=a_ctr)
    ws.row_dimensions[2].height = 13
    ws["A3"] = "Fax : +62-21-3143828, Email : corsecretary@indonesiare.co.id"
    style(ws["A3"], font=f_small, align=a_ctr)
    ws.row_dimensions[3].height = 13
    ws["A5"] = "LAPORAN KEUANGAN BUKAN KONSOLIDASI"
    style(ws["A5"], font=f_report, align=a_ctr)
    ws.row_dimensions[5].height = 18
    ws["A6"] = "Per 28 Februari 2026 dan 2025"
    style(ws["A6"], font=f_period, align=a_ctr)
    ws.row_dimensions[6].height = 15

    # Section header bars (rows 7-9)
    for rng in ["A7:D9", "J7:L9", "N7:P9"]:
        ws.merge_cells(rng)
    ws["A7"] = "LAPORAN POSISI KEUANGAN\nPER 28 Februari 2026 DAN 2025\n(Dalam Jutaan Rupiah)"
    ws["J7"] = "LAPORAN LABA RUGI DAN PENGHASILAN KOMPREHENSIF LAIN\nUntuk Tahun Yang Berakhir pada 28 Februari 2026 dan 2025\n(Dalam Jutaan Rupiah)"
    ws["N7"] = "RASIO KESEHATAN KEUANGAN\nPER 28 Februari 2026 DAN 2025\n(Dalam Jutaan Rupiah)"
    for ref in ("A7", "J7", "N7"):
        style(ws[ref], font=f_section, fill=fill_black, align=a_ctr, border=b_box)
    ws.row_dimensions[7].height = 17
    ws.row_dimensions[8].height = 13
    ws.row_dimensions[9].height = 13

    # Table column headers (row 10)
    for rng in ["A10:D10", "F10:H10", "J10:L10", "N10:P10"]:
        ws.merge_cells(rng)
    ws["A10"] = "ASET"
    ws["F10"] = "LIABILITAS DAN EKUITAS"
    ws["J10"] = "URAIAN"
    ws["N10"] = "KETERANGAN"
    for ref in ("A10", "F10", "J10", "N10"):
        style(ws[ref], font=f_table, fill=fill_grey, align=a_ctr, border=b_thin)
    ws.row_dimensions[10].height = 18

    # Year labels (row 11)
    year_labels = {
        "C11": "28 Februari 2026", "D11": "28 Februari 2025",
        "G11": "28 Februari 2026", "H11": "28 Februari 2025",
        "K11": "28 Februari 2026", "L11": "28 Februari 2025",
        "O11": "28 Februari 2026", "P11": "28 Februari 2025",
    }
    for ref, text in year_labels.items():
        ws[ref] = text
        style(ws[ref], font=f_year, fill=fill_light, align=a_ctr, border=b_thin)
    ws.row_dimensions[11].height = 15

    # Cluster content lines
    content = [ln for ln in lines if ln["y"] >= CONTENT_START_Y]
    clusters: list[dict] = []
    for ln in content:
        if not clusters or abs(ln["y"] - clusters[-1]["min_y"]) > ROW_CLUSTER_TOL:
            clusters.append({"lines": [ln], "min_y": ln["y"], "max_y": ln["y"]})
        else:
            clusters[-1]["lines"].append(ln)
            clusters[-1]["min_y"] = min(clusters[-1]["min_y"], ln["y"])
            clusters[-1]["max_y"] = max(clusters[-1]["max_y"], ln["y"])

    start_row = 12
    for idx, grp in enumerate(clusters):
        row_num = start_row + idx
        gap = (clusters[idx + 1]["min_y"] - grp["min_y"]) if idx < len(clusters) - 1 else 8.0
        ws.row_dimensions[row_num].height = max(7.0, gap * 1.15)

        parts: dict[str, list[str]] = defaultdict(list)
        for ln in grp["lines"]:
            for w in ln["words"]:
                parts[assign_col(w["x"])].append(w["text"])

        texts = [clean(" ".join(v)) for v in parts.values()]
        flags = row_flags(texts)

        for col, words in parts.items():
            val = clean(" ".join(words))
            if val:
                ws[f"{col}{row_num}"] = val

        for col in BODY_COLS:
            cell = ws[f"{col}{row_num}"]
            if cell.value is None:
                continue
            is_val = col in VALUE_COLS
            style(
                cell,
                font=f_bold if flags["bold"] else f_body,
                align=a_rgt if is_val else (a_ctr if flags["center"] else a_lft),
                border=b_thin,
                fill=fill_light if flags["bold"] and not is_val else None,
            )

    last_row = start_row + len(clusters) - 1
    ws.print_area = f"A1:P{last_row}"
    ws.freeze_panes = "A12"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
