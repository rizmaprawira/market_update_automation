"""Convert PT Reasuransi Maipark Indonesia monthly PDF to Excel (OCR-based).

Maipark publishes scanned PDFs (EPSON Scan, no text layer). This script
renders the PDF to a PNG at 300 DPI and extracts words with Tesseract TSV
output, then maps word positions to spreadsheet columns.

Usage:
    conda run -n market_update python scripts/convert_to_excel/reinsurance/convert_maipark.py

Input:  data/2026-03/raw_pdf/reasuransi/maipark/maipark_2026_03.pdf
Output: data/2026-03/raw_excel/reasuransi/maipark/maipark_2026-03.xlsx
"""
from __future__ import annotations

import csv
import re
import subprocess
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROJECT_ROOT = Path(__file__).resolve().parents[3]
PDF_PATH = PROJECT_ROOT / "data/2026-03/raw_pdf/reasuransi/maipark/maipark_2026_03.pdf"
OUT_PATH = PROJECT_ROOT / "data/2026-03/raw_excel/reasuransi/maipark/maipark_2026-03.xlsx"

OCR_DPI = 300
OCR_SCALE = 0.5
CONTENT_START_Y = 250.0
ROW_CLUSTER_TOL = 2.5

BODY_COLS = ("B", "C", "D", "F", "G", "H", "J", "K", "L", "N", "O", "P")
VALUE_COLS = {"C", "D", "G", "H", "K", "L", "O", "P"}
VALUE_COL_CENTERS = {
    "C": 370.0, "D": 445.0,
    "G": 740.0, "H": 800.0,
    "K": 1095.0, "L": 1160.0,
    "O": 1510.0, "P": 1590.0,
}

NUMERIC_RE = re.compile(r"\(?-?\d[\d,\.]*%?\)?")

BOLD_TOKENS = (
    "INVESTASI", "UTANG", "PENDAPATAN", "BEBAN", "HASIL",
    "LABA", "RUGI", "KOMISARIS", "DIREKSI", "PEMILIK PERUSAHAAN",
    "REASURADUR UTAMA", "KETERANGAN", "CATATAN", "SOLVABILITAS",
    "MMBR", "CADANGAN", "EKUITAS", "BUKAN INVESTASI",
)
CENTER_TOKENS = (
    "KETERANGAN", "SOLVABILITAS", "RASIO",
    "KOMISARIS DAN DIREKSI", "DEWAN KOMISARIS", "DIREKSI",
    "PEMILIK PERUSAHAAN", "REASURADUR UTAMA", "NAMA REASURADUR", "CATATAN",
)

OCR_REPLACEMENTS = {
    "I,": "I.", "Il,": "II.", "III,": "III.", "IV,": "IV.",
    "InvesTast": "INVESTASI", "uTANG": "UTANG",
    "IABILTTAS": "LIABILITAS", "DANEKUITAS": "DAN EKUITAS",
    "PENOAPATAN": "PENDAPATAN", "KONPREHENSIF": "KOMPREHENSIF",
    "TNOIKATOR": "INDIKATOR",
}


@dataclass(frozen=True)
class Word:
    x: float
    y: float
    text: str
    conf: float


@dataclass(frozen=True)
class Line:
    x: float
    y: float
    right: float
    words: tuple[Word, ...]


def normalize(text: str) -> str:
    text = text.replace("­", "").strip()
    return OCR_REPLACEMENTS.get(text, text)


def is_numeric(text: str) -> bool:
    return bool(NUMERIC_RE.fullmatch(text.replace(" ", "")))


def is_noise(text: str, conf: float) -> bool:
    if not text:
        return True
    if is_numeric(text):
        return False
    if conf >= 35:
        return False
    if len(text) <= 2 and text.isalpha():
        return True
    if len(text) == 1 and text in {"|", "[", "]", "(", ")", "-", "—"}:
        return True
    return False


def render_page(pdf_path: Path, tmp: Path) -> Path:
    prefix = tmp / pdf_path.stem
    subprocess.run(
        ["pdftoppm", "-r", str(OCR_DPI), "-f", "1", "-l", "1", "-png",
         str(pdf_path), str(prefix)],
        check=True,
    )
    page = tmp / f"{pdf_path.stem}-1.png"
    if not page.exists():
        raise FileNotFoundError(page)
    return page


def ocr_lines(page: Path) -> list[Line]:
    result = subprocess.run(
        ["tesseract", str(page), "stdout", "--psm", "11", "tsv"],
        check=False, capture_output=True,
    )
    if not result.stdout:
        raise RuntimeError(f"tesseract produced no output for {page.name}")

    groups: dict[tuple, list[Word]] = defaultdict(list)
    reader = csv.DictReader(result.stdout.decode("utf-8", errors="replace").splitlines(), delimiter="\t")
    for row in reader:
        if row.get("level") != "5":
            continue
        try:
            conf = float(row.get("conf") or "-1")
        except ValueError:
            conf = -1.0
        text = normalize(row.get("text") or "")
        if not text or is_noise(text, conf):
            continue
        x = float(row["left"]) * OCR_SCALE
        y = float(row["top"]) * OCR_SCALE
        groups[(row["block_num"], row["par_num"], row["line_num"])].append(
            Word(x=x, y=y, text=text, conf=conf)
        )

    lines: list[Line] = []
    for words in groups.values():
        words = sorted(words, key=lambda w: w.x)
        lines.append(Line(
            x=words[0].x,
            y=min(w.y for w in words),
            right=max(w.x for w in words),
            words=tuple(words),
        ))
    return sorted(lines, key=lambda l: (l.y, l.x))


def clean(text: str) -> str:
    text = normalize(text)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s+([,.;:%)])", r"\1", text)
    text = re.sub(r"([\(\[])\s+", r"\1", text)
    text = re.sub(r"(?<=\d),\s+(?=\d)", ",", text)
    text = re.sub(r"(?<=\d)\.\s+(?=\d)", ".", text)
    return text.strip(" |")


def is_short_int(text: str) -> bool:
    digits = re.sub(r"\D", "", text)
    return digits.isdigit() and len(digits) <= 2 and "," not in text and "." not in text and "%" not in text


def assign_col(word: Word) -> str | None:
    x = word.x
    num = is_numeric(word.text)

    if x < 960:
        if not num:
            return "B"
        if is_short_int(word.text) and x < 180:
            return "B"
        return "C" if x < 390 else "D"

    if x < 1700:
        if not num:
            return "F"
        if is_short_int(word.text) and x < 620:
            return "F"
        return "G" if x < 760 else "H"

    if x < 2400:
        if not num:
            return "J"
        if is_short_int(word.text) and x < 980:
            return "J"
        return "K" if x < 1120 else "L"

    if not num:
        return "N"
    if is_short_int(word.text) and x < 1360:
        return "N"
    return "O" if x < 1540 else "P"


def best_value(words: list[Word], col: str) -> str:
    if not words:
        return ""
    center = VALUE_COL_CENTERS.get(col, 0.0)

    def score(w: Word) -> tuple:
        t = clean(w.text)
        return (
            int(any(c in t for c in ",.%")) + int(bool(NUMERIC_RE.fullmatch(t.replace(" ", "")))),
            len(t),
            -abs(w.x - center),
            w.conf,
        )
    return clean(max(words, key=score).text)


def join_labels(words: list[Word]) -> str:
    text = " ".join(w.text for w in sorted(words, key=lambda w: w.x))
    text = clean(text)
    text = re.sub(r"\b([IVX]{1,4})\s*[,;:]?\b", lambda m: f"{m.group(1)}.", text)
    return text


def row_flags(texts: list[str]) -> dict:
    upper = " | ".join(texts).upper()
    return {
        "bold": any(t in upper for t in BOLD_TOKENS) or upper.startswith(("I.", "II.", "III.", "IV.")),
        "center": any(t in upper for t in CENTER_TOKENS),
    }


def style(cell, *, font=None, fill=None, align=None, border=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def cluster(lines: list[Line]) -> list[dict]:
    clusters: list[dict] = []
    for ln in lines:
        if not clusters or abs(ln.y - clusters[-1]["min_y"]) > ROW_CLUSTER_TOL:
            clusters.append({"lines": [ln], "min_y": ln.y, "max_y": ln.y, "min_x": ln.x, "max_x": ln.right})
        else:
            c = clusters[-1]
            c["lines"].append(ln)
            c["min_y"] = min(c["min_y"], ln.y)
            c["max_y"] = max(c["max_y"], ln.y)
            c["min_x"] = min(c["min_x"], ln.x)
            c["max_x"] = max(c["max_x"], ln.right)
    return clusters


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(PDF_PATH)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="maipark_ocr_") as tmp_dir:
        page = render_page(PDF_PATH, Path(tmp_dir))
        raw_lines = ocr_lines(page)

    wb = Workbook()
    ws = wb.active
    ws.title = "maipark"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 60
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = getattr(ws, "PAPERSIZE_A4")
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.18
    ws.page_margins.right = 0.18
    ws.page_margins.top = 0.22
    ws.page_margins.bottom = 0.22
    ws.page_margins.header = 0.08
    ws.page_margins.footer = 0.08

    col_widths = {
        "A": 4.0, "B": 32.0, "C": 12.0, "D": 12.0, "E": 2.0,
        "F": 31.0, "G": 12.0, "H": 12.0, "I": 2.0,
        "J": 40.0, "K": 12.0, "L": 12.0, "M": 2.0,
        "N": 42.0, "O": 12.0, "P": 12.0,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    f_title = Font(name="Times New Roman", size=13, bold=True)
    f_small = Font(name="Times New Roman", size=8)
    f_report = Font(name="Times New Roman", size=12, bold=True)
    f_period = Font(name="Times New Roman", size=9, bold=True)
    f_section = Font(name="Times New Roman", size=7, bold=True)
    f_table = Font(name="Times New Roman", size=7, bold=True)
    f_year = Font(name="Times New Roman", size=7, bold=True)
    f_body = Font(name="Times New Roman", size=6.5)
    f_bold = Font(name="Times New Roman", size=6.5, bold=True)
    f_tiny = Font(name="Times New Roman", size=6.1)

    a_ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
    a_lft_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    a_rgt = Alignment(horizontal="right", vertical="center", wrap_text=False)
    a_rgt_top = Alignment(horizontal="right", vertical="top", wrap_text=True)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_box = Border(left=medium, right=medium, top=medium, bottom=medium)

    fill_section = PatternFill("solid", fgColor="F0F0F0")
    fill_header = PatternFill("solid", fgColor="D9D9D9")
    fill_year = PatternFill("solid", fgColor="EFEFEF")
    fill_subtle = PatternFill("solid", fgColor="F2F2F2")

    # Header rows 1-6
    for rng in ["A1:P1", "A2:P2", "A3:P3", "A4:P4", "A5:P5", "A6:P6"]:
        ws.merge_cells(rng)
    ws["A1"] = "PT REASURANSI MAIPARK INDONESIA"
    style(ws["A1"], font=f_title, align=a_ctr)
    ws.row_dimensions[1].height = 14
    ws["A2"] = "Multivision Tower Lt.8"
    style(ws["A2"], font=f_small, align=a_ctr)
    ws.row_dimensions[2].height = 13
    ws["A3"] = "Jl. Kuningan Mulia Blok 9B, Jakarta 12960"
    style(ws["A3"], font=f_small, align=a_ctr)
    ws.row_dimensions[3].height = 13
    ws["A4"] = "Tel. (62-21) 2938 0088  Fax. (62-21) 2938 0089   E-mail: maipark@maipark.com   Website: www.maipark.com"
    style(ws["A4"], font=f_small, align=a_ctr)
    ws.row_dimensions[4].height = 18
    ws["A5"] = "LAPORAN KEUANGAN"
    style(ws["A5"], font=f_report, align=a_ctr)
    ws.row_dimensions[5].height = 18
    ws["A6"] = "Per 31 MARET 2026 dan 2025"
    style(ws["A6"], font=f_period, align=a_ctr)
    ws.row_dimensions[6].height = 15

    # Section bars (rows 7-9)
    for rng in ["A7:D9", "F7:H9", "J7:L9", "N7:P9"]:
        ws.merge_cells(rng)
    ws["A7"] = "LAPORAN POSISI KEUANGAN\nPER 31 MARET 2026 DAN 2025\n(dalam jutaan Rupiah)"
    ws["F7"] = "LIABILITAS DAN EKUITAS\nPER 31 MARET 2026 DAN 2025\n(dalam jutaan Rupiah)"
    ws["J7"] = "LAPORAN LABA (RUGI) KOMPREHENSIF\nUNTUK TAHUN YANG BERAKHIR PADA TANGGAL 31 MARET 2026 DAN 2025\n(dalam jutaan Rupiah)"
    ws["N7"] = "INDIKATOR KESEHATAN KEUANGAN\nPER 31 MARET 2026 DAN 2025\n(dalam jutaan Rupiah)"
    for ref in ("A7", "F7", "J7", "N7"):
        style(ws[ref], font=f_section, fill=fill_section, align=a_ctr, border=b_box)
    ws.row_dimensions[7].height = 17
    ws.row_dimensions[8].height = 13
    ws.row_dimensions[9].height = 13

    # Column group headers (row 10)
    for rng in ["A10:D10", "F10:H10", "J10:L10", "N10:P10"]:
        ws.merge_cells(rng)
    ws["A10"] = "ASET"
    ws["F10"] = "LIABILITAS DAN EKUITAS"
    ws["J10"] = "URAIAN"
    ws["N10"] = "URAIAN"
    for ref in ("A10", "F10", "J10", "N10"):
        style(ws[ref], font=f_table, fill=fill_header, align=a_ctr, border=b_thin)
    ws.row_dimensions[10].height = 18

    # Year labels (row 11)
    year_labels = {
        "C11": "2026", "D11": "2025",
        "G11": "2026", "H11": "2025",
        "K11": "2026", "L11": "2025",
        "O11": "2026", "P11": "2025",
    }
    for ref, text in year_labels.items():
        ws[ref] = text
        style(ws[ref], font=f_year, fill=fill_year, align=a_ctr, border=b_thin)
    ws.row_dimensions[11].height = 15

    # Content rows from OCR
    content_lines = [ln for ln in raw_lines if ln.y >= CONTENT_START_Y]
    clusters = cluster(content_lines)

    start_row = 12
    for idx, grp in enumerate(clusters):
        row_num = start_row + idx
        gap = (clusters[idx + 1]["min_y"] - grp["min_y"]) if idx < len(clusters) - 1 else 12.0
        ws.row_dimensions[row_num].height = max(7.0, min(40.0, gap * 0.55))

        min_x = grp["min_x"]
        min_y = grp["min_y"]

        # Bottom-right narrative/signature area: merge into a single wide cell
        if min_x >= 1200 and min_y >= 650:
            ws.merge_cells(f"J{row_num}:P{row_num}")
            text = clean(" ".join(w.text for ln in grp["lines"] for w in ln.words))
            if text:
                ws[f"J{row_num}"] = text
                style(ws[f"J{row_num}"], font=f_tiny,
                      align=a_rgt_top if min_y >= 980 else a_lft_top)
            continue

        parts: dict[str, list[Word]] = defaultdict(list)
        for ln in grp["lines"]:
            for w in ln.words:
                col = assign_col(w)
                if col:
                    parts[col].append(w)

        row_texts = []
        for col, words in parts.items():
            if col in VALUE_COLS:
                row_texts.append(best_value(words, col))
            else:
                row_texts.append(join_labels(words))
        flags = row_flags(row_texts)

        for col, words in parts.items():
            val = best_value(words, col) if col in VALUE_COLS else join_labels(words)
            if val:
                ws[f"{col}{row_num}"] = val

        for col in BODY_COLS:
            cell = ws[f"{col}{row_num}"]
            if cell.value is None:
                continue
            is_val = col in VALUE_COLS
            style(
                cell,
                font=f_bold if flags["bold"] else f_body,
                align=a_rgt if is_val else (a_ctr if flags["center"] else a_lft),
                border=b_thin,
                fill=fill_subtle if flags["bold"] and not is_val else None,
            )

    last_row = start_row + len(clusters) - 1
    ws.print_area = f"A1:P{last_row}"
    ws.freeze_panes = "A12"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
