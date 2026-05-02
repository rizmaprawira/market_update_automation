"""Convert PT Maskapai Reasuransi Indonesia Tbk. (MAREIN) monthly PDF to Excel.

Marein uses a bilingual (Indonesian/English) layout with 6 data zones across the page:
  B: Indonesian asset labels     C: asset value Mar-26   D: asset value Dec-25
  F: English asset labels        G: Indonesian liab labels
  H: liab value Mar-26           J: liab value Dec-25    K: English liab labels
  L: Income statement labels     N: income value Mar-26  O: income value Mar-25
  P: English income labels

Usage:
    conda run -n market_update python scripts/convert_to_excel/reinsurance/convert_marein.py

Input:  data/2026-03/raw_pdf/reasuransi/marein/marein_2026_03.pdf
Output: data/2026-03/raw_excel/reasuransi/marein/marein_2026-03.xlsx
"""
from __future__ import annotations

import subprocess
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROJECT_ROOT = Path(__file__).resolve().parents[3]
PDF_PATH = PROJECT_ROOT / "data/2026-03/raw_pdf/reasuransi/marein/marein_2026_03.pdf"
OUT_PATH = PROJECT_ROOT / "data/2026-03/raw_excel/reasuransi/marein/marein_2026-03.xlsx"

CONTENT_START_Y = 117.0
ROW_CLUSTER_TOL = 0.8

BODY_COLS = ("B", "C", "D", "F", "G", "H", "J", "K", "L", "N", "O", "P")
VALUE_COLS = {"C", "D", "H", "J", "N", "O"}
LABEL_COLS = {"B", "F", "G", "K", "L", "P"}

# A4 landscape (841.92 × 595.32). Calibrated from 2026-03 PDF.
# B < 135: Indonesian asset labels (19-115) and row numbers (19-24)
# C 135-165: asset value Mar-26 (141-152)
# D 165-228: asset value Dec-25 (173-183)
# F 228-285: English asset labels (229-280)
# G 285-400: Indonesian liability labels (287-335)
# H 400-430: liability value Mar-26 (404-415)
# J 430-470: liability value Dec-25 (437-447)
# K 470-510: English liability labels (472-506)
# L 510-635: Income statement labels (512-610)
# N 635-670: income value Mar-26 (654-664)
# O 670-715: income value Mar-25 (686-697)
# P 715+: English income labels (748-793)
X_THRESHOLDS = (135, 165, 228, 285, 400, 430, 470, 510, 635, 670, 715)

BOLD_TOKENS = (
    "JUMLAH", "TOTAL", "LABA", "RUGI", "HASIL",
    "INVESTASI", "BUKAN INVESTASI",
    "KOMISARIS", "DIREKSI", "PEMILIK PERUSAHAAN",
    "REASURADUR UTAMA", "KETERANGAN",
    "PENDAPATAN UNDERWRITING", "BEBAN UNDERWRITING",
    "LABA (RUGI)", "TOTAL EQUITY", "TOTAL LIABILITIES",
)
CENTER_TOKENS = (
    "KETERANGAN", "URAIAN", "REASURADUR",
    "NAMA REASURADUR", "CATATAN",
)


def parse_words(pdf_path: Path) -> list[dict]:
    xml = subprocess.check_output(["pdftotext", "-bbox-layout", str(pdf_path), "-"], text=True)
    root = ET.fromstring(xml)
    ns = {"x": "http://www.w3.org/1999/xhtml"}
    lines = []
    for line in root.findall(".//x:line", ns):
        words = []
        for w in line.findall("x:word", ns):
            text = (w.text or "").strip()
            if text:
                words.append({"x": float(w.attrib["xMin"]), "text": text})
        if words:
            words.sort(key=lambda ww: ww["x"])
            lines.append({"y": float(line.attrib["yMin"]), "words": words})
    lines.sort(key=lambda ll: (ll["y"], ll["words"][0]["x"]))
    return lines


def assign_col(x: float) -> str:
    for thresh, col in zip(X_THRESHOLDS, BODY_COLS[:-1]):
        if x < thresh:
            return col
    return BODY_COLS[-1]


def clean(text: str) -> str:
    return " ".join(text.split())


def row_flags(texts: list[str]) -> dict:
    joined = " | ".join(texts).upper()
    return {
        "bold": any(t in joined for t in BOLD_TOKENS)
                or joined.startswith(("I.", "II.", "III.", "IV.")),
        "center": any(t in joined for t in CENTER_TOKENS),
    }


def style(cell, *, font=None, fill=None, align=None, border=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(PDF_PATH)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    lines = parse_words(PDF_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "marein"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 60
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = getattr(ws, "PAPERSIZE_A4")
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.18
    ws.page_margins.right = 0.18
    ws.page_margins.top = 0.22
    ws.page_margins.bottom = 0.22
    ws.page_margins.header = 0.08
    ws.page_margins.footer = 0.08

    col_widths = {
        "A": 3.0, "B": 26.0, "C": 10.0, "D": 10.0, "E": 1.5,
        "F": 22.0, "G": 22.0, "H": 10.0, "I": 1.5,
        "J": 10.0, "K": 22.0, "L": 1.5,
        "M": 40.0, "N": 10.0, "O": 10.0, "P": 26.0,
    }
    # Non-standard column layout for marein: uses M as income label (not L)
    # and L as spacer. Remap after standard setup.
    col_widths = {
        "A": 3.0, "B": 26.0, "C": 10.5, "D": 10.5, "E": 1.5,
        "F": 22.0, "G": 22.0, "H": 10.5, "I": 1.5,
        "J": 10.5, "K": 22.0, "L": 2.0,
        "M": 2.0, "N": 10.5, "O": 10.5, "P": 26.0,
    }
    # For L (income labels), we'll actually use the extracted content in the L slot
    # but rename the column heading. The content extraction still uses BODY_COLS.
    col_widths["L"] = 42.0  # income labels column (mapped from PDF x 510-635)
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    f_title = Font(name="Times New Roman", size=11, bold=True)
    f_small = Font(name="Times New Roman", size=8)
    f_report = Font(name="Times New Roman", size=10, bold=True)
    f_period = Font(name="Times New Roman", size=8, bold=True)
    f_section = Font(name="Times New Roman", size=6.5, bold=True, color="FFFFFF")
    f_table = Font(name="Times New Roman", size=6.5, bold=True)
    f_year = Font(name="Times New Roman", size=6.5, bold=True)
    f_body = Font(name="Times New Roman", size=6.2)
    f_bold = Font(name="Times New Roman", size=6.2, bold=True)
    f_en = Font(name="Times New Roman", size=5.8, italic=True)

    a_ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
    a_rgt = Alignment(horizontal="right", vertical="center", wrap_text=False)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_box = Border(left=medium, right=medium, top=medium, bottom=medium)

    fill_section = PatternFill("solid", fgColor="1F3763")
    fill_header = PatternFill("solid", fgColor="D9E2F3")
    fill_year = PatternFill("solid", fgColor="EEF3FB")
    fill_subtle = PatternFill("solid", fgColor="F2F2F2")
    fill_en = PatternFill("solid", fgColor="F8F8FF")

    # Header rows 1-6
    for rng in ["A1:P1", "A2:P2", "A3:P3", "A5:P5", "A6:P6"]:
        ws.merge_cells(rng)
    ws["A1"] = "RINGKASAN LAPORAN KEUANGAN / FINANCIAL STATEMENTS SUMMARY"
    style(ws["A1"], font=f_title, align=a_ctr)
    ws.row_dimensions[1].height = 16
    ws["A2"] = "PT MASKAPAI REASURANSI INDONESIA TBK."
    style(ws["A2"], font=Font(name="Times New Roman", size=11, bold=True), align=a_ctr)
    ws.row_dimensions[2].height = 15
    ws["A3"] = "LAPORAN BULANAN 31 MARET 2026 / MONTHLY REPORT MARCH 31, 2026"
    style(ws["A3"], font=f_small, align=a_ctr)
    ws.row_dimensions[3].height = 13
    ws["A5"] = "LAPORAN POSISI KEUANGAN / STATEMENTS OF FINANCIAL POSITION"
    style(ws["A5"], font=f_report, align=a_ctr)
    ws.row_dimensions[5].height = 18
    ws["A6"] = "PER 31 MARET 2026 DAN 31 DESEMBER 2025 / AS OF MARCH 31, 2026 AND DECEMBER 31, 2025"
    style(ws["A6"], font=f_period, align=a_ctr)
    ws.row_dimensions[6].height = 15

    # Section bars (rows 7-9): Balance sheet left | Income statement right
    ws.merge_cells("A7:K9")
    ws["A7"] = "BALANCE SHEET / LAPORAN POSISI KEUANGAN\n(dalam jutaan rupiah / in million rupiah)"
    style(ws["A7"], font=f_section, fill=fill_section, align=a_ctr, border=b_box)
    ws.merge_cells("L7:P9")
    ws["L7"] = "INCOME STATEMENT / LAPORAN LABA RUGI\nUNTUK TAHUN YANG BERAKHIR / FOR THE YEARS ENDED 31 MARET 2026 DAN 2025\n(dalam jutaan rupiah / in million rupiah)"
    style(ws["L7"], font=f_section, fill=fill_section, align=a_ctr, border=b_box)
    ws.row_dimensions[7].height = 17
    ws.row_dimensions[8].height = 13
    ws.row_dimensions[9].height = 13

    # Column group headers (row 10)
    ws.merge_cells("A10:D10")
    ws["A10"] = "ASET / ASSETS"
    style(ws["A10"], font=f_table, fill=fill_header, align=a_ctr, border=b_thin)
    ws.merge_cells("F10:K10")
    ws["F10"] = "LIABILITAS DAN EKUITAS / LIABILITIES AND EQUITY"
    style(ws["F10"], font=f_table, fill=fill_header, align=a_ctr, border=b_thin)
    ws.merge_cells("L10:P10")
    ws["L10"] = "LAPORAN LABA RUGI / INCOME STATEMENT"
    style(ws["L10"], font=f_table, fill=fill_header, align=a_ctr, border=b_thin)
    ws.row_dimensions[10].height = 18

    # Year labels (row 11)
    year_cells = {
        "C11": "31 Mar 2026", "D11": "31 Des 2025",
        "H11": "31 Mar 2026", "J11": "31 Des 2025",
        "N11": "31 Mar 2026", "O11": "31 Mar 2025",
    }
    desc_cells = {
        "F11": "Keterangan / Description",
        "G11": "Indonesian Liab. / Liabilitas",
        "K11": "English / Inggris",
        "P11": "Description / Keterangan",
    }
    for ref, text in {**year_cells, **desc_cells}.items():
        ws[ref] = text
        if ref in year_cells:
            style(ws[ref], font=f_year, fill=fill_year, align=a_ctr, border=b_thin)
        else:
            style(ws[ref], font=Font(name="Times New Roman", size=6, italic=True),
                  fill=fill_en, align=a_ctr, border=b_thin)
    ws.row_dimensions[11].height = 15

    # Extract content
    content = [ln for ln in lines if ln["y"] >= CONTENT_START_Y]
    clusters: list[dict] = []
    for ln in content:
        if not clusters or abs(ln["y"] - clusters[-1]["min_y"]) > ROW_CLUSTER_TOL:
            clusters.append({"lines": [ln], "min_y": ln["y"], "max_y": ln["y"]})
        else:
            clusters[-1]["lines"].append(ln)
            clusters[-1]["min_y"] = min(clusters[-1]["min_y"], ln["y"])
            clusters[-1]["max_y"] = max(clusters[-1]["max_y"], ln["y"])

    start_row = 12
    for idx, grp in enumerate(clusters):
        row_num = start_row + idx
        gap = (clusters[idx + 1]["min_y"] - grp["min_y"]) if idx < len(clusters) - 1 else 8.0
        ws.row_dimensions[row_num].height = max(7.0, min(35.0, gap * 0.9))

        parts: dict[str, list[str]] = defaultdict(list)
        for ln in grp["lines"]:
            for w in ln["words"]:
                parts[assign_col(w["x"])].append(w["text"])

        texts = [clean(" ".join(v)) for v in parts.values()]
        flags = row_flags(texts)

        for col, words in parts.items():
            val = clean(" ".join(words))
            if val:
                ws[f"{col}{row_num}"] = val

        for col in BODY_COLS:
            cell = ws[f"{col}{row_num}"]
            if cell.value is None:
                continue
            is_val = col in VALUE_COLS
            is_en = col in {"F", "K", "P"}
            style(
                cell,
                font=f_bold if flags["bold"] and not is_en else (f_en if is_en else f_body),
                align=a_rgt if is_val else (a_ctr if flags["center"] else a_lft),
                border=b_thin,
                fill=fill_en if is_en else (fill_subtle if flags["bold"] and not is_val else None),
            )

    last_row = start_row + len(clusters) - 1
    ws.print_area = f"A1:P{last_row}"
    ws.freeze_panes = "A12"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
