"""Fill the OJK standardized template with values extracted from a company's PDF.

Strategy:
  1. Copy the template file as base (all styling, merged cells preserved).
  2. Fill company header + period in rows 1-6.
  3. For each extracted row, identify which template row the label matches
     and write the numeric values into the correct value columns.

Template value columns:
  Assets      → G (current), H (prior)     | labels in column B, rows 12-61
  Liabilities → O (current), P (prior)     | labels in column J
  Income      → X (current), Y (prior)     | labels in column S
  Health      → AF (current), AG (prior)   | labels in column AA
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from copy import copy

TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "laporan_keuangan_template_reasuransi.xlsx"

_VAL_COLS = {
    "assets":      ("G",  "H"),
    "liabilities": ("O",  "P"),
    "income":      ("X",  "Y"),
    "health":      ("AF", "AG"),
}
_LABEL_COLS = {
    "assets":      "B",
    "liabilities": "J",
    "income":      "S",
    "health":      "AA",
}
_DATA_ROWS = range(12, 62)

# Synonym replacements applied BEFORE label matching (both sides normalised)
_SYNONYMS = {
    "retrosesi": "reasuransi",  # reinsurers say retrosesi; template says reasuransi
    "kewajiban": "liabilitas",
    "hutang":    "utang",
    "piutang":   "tagihan",
    "risikio":   "risiko",      # common PDF typo
    "reksadana": "reksa dana",  # some PDFs write as one word
    "netto":     "neto",        # variant spelling
}


# ── Normalisation ────────────────────────────────────────────────────────────

def _norm(text: str) -> str:
    """Lowercase, strip item-number prefixes + parentheticals, apply synonyms."""
    if not text:
        return ""
    t = str(text).strip()
    # Strip leading roman numerals: I, II, III, IV, V, VI ...
    t = re.sub(r"^(IV|IX|V?I{1,3}|XI{0,2})\s*\.?\s+", "", t, flags=re.IGNORECASE)
    # Strip leading arabic item numbers:  "1.", "21.", "33 "
    t = re.sub(r"^\d+\s*\.?\s+", "", t)
    # Strip leading sub-item letters: "a.", "b.", "c.", "A.", "B."
    t = re.sub(r"^[a-eA-E]\s*\.?\s+", "", t)
    # Strip ALL parentheticals (formula refs AND semantic ones)
    t = re.sub(r"\s*\([^)]*\)", "", t)
    # Strip asterisks (template uses *) as footnote markers)
    t = t.replace("*", "")
    t = " ".join(t.split()).lower()
    for src, dst in _SYNONYMS.items():
        t = t.replace(src, dst)
    return t


# ── Value parsing ────────────────────────────────────────────────────────────

_NUM_RE = re.compile(r"^\(?-?[\d.,]+\)?%?$")


def _parse_val(text: str, thousands_sep: str = ".") -> float | None:
    """Parse Indonesian or US number format → float.

    Indonesian (thousands_sep="."):
      "1.234.567"     → 1234567.0   (dots = thousands)
      "1.234.567,89"  → 1234567.89  (comma = decimal)

    US (thousands_sep=","):
      "1,234,567"     → 1234567.0   (commas = thousands)
      "1,234,567.89"  → 1234567.89  (dot = decimal)

    Both:
      "(2.345)"       → -2345.0     (parentheses = negative)
      "93,3%"         → 93.3        (percentage; caller decides semantics)
      "-"             → None
    """
    if not text:
        return None
    t = str(text).strip()
    if t in ("-", ""):
        return None
    negative = t.startswith("(") and t.endswith(")")
    if negative:
        t = t[1:-1]
    t = t.rstrip("%").replace(" ", "")
    if not t:
        return None

    decimal_sep = "," if thousands_sep == "." else "."

    if decimal_sep in t:
        # has a decimal part
        parts = t.rsplit(decimal_sep, 1)
        try:
            val = float(parts[0].replace(thousands_sep, "") + "." + parts[1])
        except (ValueError, IndexError):
            return None
    else:
        try:
            val = float(t.replace(thousands_sep, ""))
        except ValueError:
            return None
    return -val if negative else val


def _split_two(s: str) -> tuple[str, str]:
    """Split a string that contains two space-separated numbers."""
    if not s:
        return "", ""
    parts = s.split()
    if len(parts) >= 2 and _NUM_RE.match(parts[0]) and _NUM_RE.match(parts[-1]):
        return parts[0], parts[-1]
    return s, ""


# ── Template index ───────────────────────────────────────────────────────────

def _build_index(ws) -> dict[str, dict[str, int]]:
    """Build {section → {normalized_label → row_number}} from the template sheet."""
    idx: dict[str, dict[str, int]] = {s: {} for s in _LABEL_COLS}
    for row in _DATA_ROWS:
        for sec, col in _LABEL_COLS.items():
            v = ws[f"{col}{row}"].value
            if v:
                n = _norm(str(v))
                if n and n not in idx[sec]:   # first occurrence wins
                    idx[sec][n] = row
    return idx


def _find_row(
    label: str,
    section: str,
    idx: dict[str, dict[str, int]],
    ovr: dict[str, dict[str, int]],
) -> int | None:
    n = _norm(label)
    if not n:
        return None
    # 1. Per-company override (keys already normalised)
    if n in ovr.get(section, {}):
        return ovr[section][n]
    # 2. Exact normalised match
    if n in idx[section]:
        return idx[section][n]
    # 3. Extracted label is contained in template label
    for t, row in idx[section].items():
        if n in t:
            return row
    # 4. Template label is contained in extracted label
    for t, row in idx[section].items():
        if t in n:
            return row
    return None


# ── Bottom section (rows 62-75): komisaris, direksi, pemilik, reasuransi ─────

# Role keyword → category (ordered most-specific first)
_ROLE_MAP = [
    ("komisaris_utama", ["KOMISARIS UTAMA", "PRESIDEN KOMISARIS", "KOMISARIS INDEPENDEN DAN UTAMA",
                         "UTAMA/INDEPENDEN", "UTAMA INDEPENDEN", "KOMISARIS UTAMA/INDEPENDEN"]),
    ("direktur_utama",  ["PLT. DIREKTUR UTAMA", "PLT DIREKTUR UTAMA", "DIREKTUR UTAMA",
                         "PRESIDEN DIREKTUR", "PRESIDENT DIRECTOR"]),
    ("komisaris",       ["KOMISARIS INDEPENDEN", "KOMISARIS NON-INDEPENDEN", "KOMISARIS NON INDEPENDEN",
                         "KOMISARIS"]),
    ("direktur_teknik", ["DIREKTUR TEKNIK"]),
    ("direktur_aktuari",["DIREKTUR AKTUARI"]),
    ("direktur",        ["WAKIL DIREKTUR UTAMA", "WAKIL DIREKTUR", "DIREKTUR", "PLT. DIREKTUR"]),
]

def _classify_role(text: str) -> str | None:
    tu = text.upper()
    for role_key, keywords in _ROLE_MAP:
        if any(kw in tu for kw in keywords):
            return role_key
    return None


def _extract_name(text: str) -> str | None:
    """Pull person name from text formatted as one of:
      'ROLE : Name'        →  Name
      'ROLE Name :'        →  Name  (name between role keyword and trailing colon)
      ': Name ROLE ...'    →  Name  (leading colon, role suffix)
      'Name ROLE ...'      →  Name  (role appended after name, no colon)
    """
    t = text.strip()

    # Leading colon pattern: ": Name ROLE"
    if t.startswith(":"):
        rest = t[1:].strip()
        # Strip role suffix
        rest_u = rest.upper()
        for _, keywords in _ROLE_MAP:
            for kw in keywords:
                idx = rest_u.find(kw)
                if idx != -1:
                    candidate = rest[:idx].strip().rstrip(",")
                    if candidate:
                        return candidate
        return rest or None

    # " : " separator: "ROLE : Name"
    if " : " in t:
        parts = t.split(" : ", 1)
        return parts[1].strip() or None

    # Trailing colon: "ROLE Name :"  →  strip role keyword, then trailing colon
    if t.rstrip().endswith(":"):
        t_clean = t.rstrip(":").strip()
        t_upper = t_clean.upper()
        for _, keywords in _ROLE_MAP:
            for kw in keywords:
                idx = t_upper.find(kw)
                if idx != -1:
                    candidate = t_clean[idx + len(kw):].strip()
                    if candidate:
                        return candidate
        return None

    # "Name ROLE" — role keyword appended to name (marein G column style)
    # Find the EARLIEST role keyword in the string to avoid returning partial role text
    t_upper = t.upper()
    earliest_idx = len(t)
    for _, keywords in _ROLE_MAP:
        for kw in keywords:
            pos = t_upper.find(kw)
            if 0 < pos < earliest_idx:
                earliest_idx = pos
    if earliest_idx < len(t):
        candidate = t[:earliest_idx].strip().rstrip(",")
        if candidate:
            return candidate

    return None


def _copy_cell_style(source_cell, target_cell) -> None:
    """Copy formatting (font, fill, border, alignment) from source to target cell."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = copy(source_cell.number_format)


def _get_writable_cell(ws, col: str, row: int):
    """Get the cell to write to, handling merged cells.

    If the cell is part of a merged range (MergedCell), return the top-left cell
    of that range. Otherwise, return the cell itself.
    """
    cell = ws[f"{col}{row}"]
    # Check if this cell is a MergedCell (part of a merged range)
    if cell.__class__.__name__ == 'MergedCell':
        # Find which merged range this cell belongs to
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= cell.column <= merged_range.max_col and \
               merged_range.min_row <= cell.row <= merged_range.max_row:
                # Return the top-left cell of this merged range
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def fill_bottom_section(ws, rows: list[dict], config: dict) -> None:
    """Fill template bottom section with komisaris, direksi, pemilik, reasuransi data.

    Dynamically extends rows as needed to accommodate variable numbers of entries:
    - Directors (komisaris, direktur) expanded beyond template rows 64-70
    - Owners (pemilik) extended beyond template rows 72-75
    - Reinsurers dalam/luar extended beyond template rows 65-68 / 71-74

    Director subtypes are captured: direktur_teknik, direktur_aktuari, plus general direktur.
    """
    bs = config.get("bottom_section", {})
    if not bs:
        return

    kom_col     = bs.get("komisaris_col", "B")
    pem_col     = bs.get("pemilik_col", kom_col)
    pem_pct_col = bs.get("pemilik_pct_col")
    re_col      = bs.get("reasuransi_col")
    re_pct_col  = bs.get("reasuransi_pct_col")
    jkt_col     = bs.get("jakarta_col", "N")
    re_default  = bs.get("reasuransi_default_zone", None)  # 'luar' or 'dalam'

    # ── Collect data ──────────────────────────────────────────────────────────
    komisaris_utama: str | None = None
    komisaris_list:  list[str]  = []
    direktur_utama:  str | None = None
    direktur_teknik: list[str]  = []
    direktur_aktuari:list[str]  = []
    direktur_list:   list[str]  = []
    pemilik_list:    list[tuple[str, str]] = []   # (name, pct)
    re_dalam:        list[tuple[str, str]] = []   # (name, pct)
    re_luar:         list[tuple[str, str]] = []   # (name, pct)
    jakarta_date:    str | None = None

    state      = None   # 'komisaris' | 'direksi' | 'pemilik'
    prev_role  = None

    # ── Pass 1: komisaris / direksi / pemilik ─────────────────────────────────
    for row in rows:
        # Jakarta date (from configured column)
        jt = str(row.get(jkt_col, "") or "").strip()
        if jt.startswith("Jakarta") and not jakarta_date:
            jakarta_date = jt

        t  = str(row.get(kom_col, "") or "").strip()
        if not t:
            continue
        tu = t.upper()

        # Section header transitions
        if any(m in tu for m in ("DEWAN KOMISARIS", "KOMISARIS DAN DIREKSI")):
            state = "komisaris"
            continue
        if tu in ("DIREKSI", "DIREKSI :", "DIREKSI:"):
            state = "direksi"
            continue
        if "PEMILIK PERUSAHAAN" in tu:
            state = "pemilik"
            continue

        # Section terminators — stop collecting into any state
        if any(m in tu for m in ("DEWAN PENGAWAS SYARIAH", "ANAK PERUSAHAAN")) or tu.startswith("@"):
            state = None
            continue

        # ── Pemilik entries ───────────────────────────────────────────────────
        if state == "pemilik" and pem_col == kom_col:
            entry = re.sub(r"^\d+[\s.]+", "", t)  # strip leading "1.", "2 "
            # Check dedicated pct column first (e.g. indonesiare D col)
            col_pct = str(row.get(pem_pct_col, "") or "").strip() if pem_pct_col else ""
            if col_pct:
                pemilik_list.append((entry, col_pct))
            elif " : " in entry:
                name, _, pct = entry.partition(" : ")
                pemilik_list.append((name.strip(), pct.strip()))
            else:
                # percentage embedded: "Name pct%" or "Name : pct"
                m = re.search(r"^(.*?)\s+(\d+[\.,]\d+%)\s*$", entry)
                if m:
                    pemilik_list.append((m.group(1).strip(), m.group(2).strip()))
                elif re.search(r"\d+%", entry):
                    # "Name 50%"
                    m2 = re.search(r"^(.*?)\s+(\d+%)\s*$", entry)
                    if m2:
                        pemilik_list.append((m2.group(1).strip(), m2.group(2).strip()))
                    else:
                        pemilik_list.append((entry, ""))
                else:
                    pemilik_list.append((entry, ""))
            continue

        # ── Komisaris / direksi entries ───────────────────────────────────────
        if state in ("komisaris", "direksi", None):
            # Continuation line (name on its own line)
            if t.startswith(":"):
                name = t[1:].strip()
                # might contain role at the end: ": Name ROLE"
                role_in_cont = _classify_role(name)
                if role_in_cont:
                    # reversed format: name is before role keyword
                    extracted = _extract_name(t)
                    if extracted:
                        name = extracted
                        prev_role = role_in_cont
                if name:
                    if prev_role == "komisaris_utama" and not komisaris_utama:
                        komisaris_utama = name
                    elif prev_role == "komisaris":
                        komisaris_list.append(name)
                    elif prev_role == "direktur_utama" and not direktur_utama:
                        direktur_utama = name
                    elif prev_role == "direktur_teknik":
                        direktur_teknik.append(name)
                    elif prev_role == "direktur_aktuari":
                        direktur_aktuari.append(name)
                    elif prev_role == "direktur":
                        direktur_list.append(name)
                continue

            role_key = _classify_role(t)
            if role_key:
                prev_role = role_key
                name = _extract_name(t)
                if name:
                    if role_key == "komisaris_utama" and not komisaris_utama:
                        komisaris_utama = name
                    elif role_key == "komisaris":
                        komisaris_list.append(name)
                    elif role_key == "direktur_utama" and not direktur_utama:
                        direktur_utama = name
                    elif role_key == "direktur_teknik":
                        direktur_teknik.append(name)
                    elif role_key == "direktur_aktuari":
                        direktur_aktuari.append(name)
                    elif role_key == "direktur":
                        direktur_list.append(name)

    # ── Pass 1b: pemilik from separate column (e.g. tugure F column) ─────────
    if pem_col != kom_col:
        in_pem = False
        for row in rows:
            pt = str(row.get(pem_col, "") or "").strip()
            if not pt:
                continue
            ptu = pt.upper()
            if "PEMILIK PERUSAHAAN" in ptu:
                in_pem = True
                continue
            if not in_pem:
                continue
            # Skip pure headers / section markers
            if any(skip in ptu for skip in ("REASURADUR", "NAMA REASURADUR", "DIREKSI",
                                             "KOMISARIS", "KETERANGAN", "%", "URAIAN")):
                continue
            entry = re.sub(r"^\d+[\s.]+", "", pt)
            pct_str = str(row.get(pem_pct_col, "") or "").strip() if pem_pct_col else ""
            if not pct_str:
                # Try to extract pct from entry text
                m = re.search(r"(\d+[.,]\d+%|\d+%)", entry)
                if m:
                    pct_str = m.group(1)
                    entry = entry[:entry.rfind(m.group(1))].strip()
            if entry:
                pemilik_list.append((entry, pct_str))

    # ── Pass 2: reasuransi ────────────────────────────────────────────────────
    if re_col:
        re_state = None  # only start collecting AFTER the REASURADUR UTAMA header
        re_section_found = False
        for row in rows:
            rt = str(row.get(re_col, "") or "").strip()
            if not rt or rt in ("-", "%"):
                continue
            rtu = rt.upper()

            # Section header: activate collection zone
            if "REASURADUR UTAMA" in rtu or "NAMA REASURADUR" in rtu:
                re_section_found = True
                if re_state is None:
                    re_state = re_default   # apply default zone (luar/dalam/None)
                continue

            # Skip other non-data header lines
            if any(skip in rtu for skip in ("KETERANGAN", "DIREKSI", "PT MASKAPAI",
                                             "S.E. & O", "S.E &", "BOARD OF")):
                continue

            # Wait until section header seen before collecting anything
            if not re_section_found:
                continue

            # Zone transitions
            if any(m in rtu for m in ("DALAM NEGERI", "REASURANSI JIWA",
                                       "I. REASURANSI", "A. REASURANSI")):
                re_state = "dalam"
                continue
            if any(m in rtu for m in ("LUAR NEGERI", "REASURANSI UMUM",
                                       "II. REASURANSI", "B. REASURANSI",
                                       "REASURADUR LUAR NEGERI")):
                re_state = "luar"
                continue
            if re_state is None:
                continue

            # Get percentage
            pct_str = str(row.get(re_pct_col, "") or "").strip() if re_pct_col else ""
            # Strip leading number from name
            name = re.sub(r"^\d+[\s.]+", "", rt).strip()
            # Try "Name PCT%" embedded in cell
            m = re.match(r"^(.*?)\s+(\d+[.,]\d+%)\s*(?:\d+\.\s*)?$", name)
            if m:
                name, pct_str = m.group(1).strip(), m.group(2).strip()

            if name and len(name) > 1 and name not in ("-", "OTHERS"):
                if re_state == "dalam":
                    re_dalam.append((name, pct_str))
                else:
                    re_luar.append((name, pct_str))

    # ── Write to template ─────────────────────────────────────────────────────

    # Calculate end row for each section (accounting for actual data size)
    # Commissioners: E64 (utama) + len(komisaris) + space for direktur
    dir_other = len(direktur_teknik) + len(direktur_aktuari) + len(direktur_list)
    dir_end = 68 + max(2, dir_other)  # direktur_utama at 68, min 2 more rows, more if needed

    # Reinsurers dalam: start at J65, end based on data
    dalam_end = 65 + len(re_dalam) - 1

    # Reinsurers luar: always start at row 71 (template allocated) to overwrite placeholders
    luar_template_start = 71
    luar_end = luar_template_start + len(re_luar) - 1

    # Owners start after reinsurers luar (never overlap)
    pemilik_start = luar_end + 1

    # Helper to write cell and copy style from template row if needed
    def write_with_style(col: str, row: int, value, source_row: int) -> None:
        cell = _get_writable_cell(ws, col, row)
        cell.value = value
        # Copy style from source if row extends beyond template (row > max template row for this section)
        if row > source_row:
            source_cell = _get_writable_cell(ws, col, source_row)
            _copy_cell_style(source_cell, cell)

    # Write commissioners to column E
    if komisaris_utama:
        cell = _get_writable_cell(ws, "E", 64)
        cell.value = komisaris_utama
    for i, name in enumerate(komisaris_list):
        write_with_style("E", 65+i, name, 65)

    if direktur_utama:
        cell = _get_writable_cell(ws, "E", 68)
        cell.value = direktur_utama

    # Write all other directors: teknik, aktuari, then general
    dir_row = 69
    source_row = 69  # Row to copy style from
    for name in direktur_teknik:
        write_with_style("E", dir_row, name, source_row)
        source_row = dir_row
        dir_row += 1
    for name in direktur_aktuari:
        write_with_style("E", dir_row, name, source_row)
        source_row = dir_row
        dir_row += 1
    for name in direktur_list:
        write_with_style("E", dir_row, name, source_row)
        source_row = dir_row
        dir_row += 1

    # Write reinsurers dalam (column J)
    # Template allocates rows 65-68 (4 slots), extend beyond if needed
    for i, (name, pct) in enumerate(re_dalam):
        write_with_style("J", 65+i, f"{i+1}. {name}", 65)
        if pct:
            write_with_style("O", 65+i, pct, 65)

    # Write reinsurers luar (column J)
    # Template allocates rows 71-74 (4 slots), but always start from 71 to overwrite placeholders
    # If more than 4, extend beyond row 74
    luar_template_start = 71
    for i, (name, pct) in enumerate(re_luar):
        write_with_style("J", luar_template_start+i, f"{i+1}. {name}", 71)
        if pct:
            write_with_style("O", luar_template_start+i, pct, 71)

    # Write owners (column A)
    for i, (name, pct) in enumerate(pemilik_list):
        write_with_style("A", pemilik_start+i, f"{i+1}. {name}", 72)
        if pct:
            write_with_style("F", pemilik_start+i, pct, 72)

    if jakarta_date:
        cell = _get_writable_cell(ws, "AA", 62)
        cell.value = jakarta_date


# ── Main fill function ───────────────────────────────────────────────────────

def fill_template(
    rows: list[dict],
    config: dict,
    period: str,
    out_path: Path,
) -> None:
    """Copy template, write company header + extracted numbers, save to out_path."""
    # Import here to avoid circular imports
    from common.template_writer import compute_dates, _fmt

    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE_PATH, out_path)

    wb = load_workbook(out_path)
    ws = wb.active
    dates = compute_dates(period, config)

    # ── Header rows 1-6 ──────────────────────────────────────────────────────
    hr = config.get("header_rows", {})
    ws["A1"] = _fmt(hr.get(1, hr.get("1", {})).get("text", ""), dates)
    ws["A2"] = _fmt(hr.get(2, hr.get("2", {})).get("text", ""), dates)
    ws["A3"] = _fmt(hr.get(3, hr.get("3", {})).get("text", ""), dates)
    ws["A6"] = _fmt(hr.get(6, hr.get("6", {})).get("text", "Per {day} {month_full} {YEAR4} dan {PRIOR_YEAR4}"), dates)

    # ── Year column headers (row 10) ─────────────────────────────────────────
    pcfg = config.get("period", {})
    lbl_cur = _fmt(pcfg.get("current_label", "{YEAR4}"),       dates)
    lbl_pri = _fmt(pcfg.get("prior_label",   "{PRIOR_YEAR4}"), dates)
    for col in ("G", "O", "X", "AF"):
        ws[f"{col}10"] = lbl_cur
    for col in ("H", "P", "Y", "AG"):
        ws[f"{col}10"] = lbl_pri

    # ── Build label index + normalised overrides ──────────────────────────────
    idx = _build_index(ws)
    raw_ovr = config.get("label_map", {})
    ovr: dict[str, dict[str, int]] = {
        sec: {_norm(k): v for k, v in mapping.items()}
        for sec, mapping in raw_ovr.items()
    }

    # ── Zone config: which extracted columns hold label / current / prior ──────
    #    Defaults work for all standard companies (nasre, tugure, orionre,
    #    indonesiare, inare).  Override in configs/marein.yml.
    defaults = {
        "assets":      ("B", "C", "D"),
        "liabilities": ("F", "G", "H"),
        "income":      ("J", "K", "L"),
        "health":      ("N", "O", "P"),
    }
    zcfg = config.get("zones", {})
    zones: dict[str, tuple[str, str, str] | None] = {}
    for sec, (dl, dc, dp) in defaults.items():
        zs = zcfg.get(sec)
        if zs is False or zs == "false":
            zones[sec] = None               # section absent in this company's PDF
        elif isinstance(zs, dict):
            zones[sec] = (zs.get("label_col", dl),
                          zs.get("val_current", dc),
                          zs.get("val_prior",   dp))
        else:
            zones[sec] = (dl, dc, dp)       # use defaults

    # ── Fill data rows ────────────────────────────────────────────────────────
    tsep = config.get("thousands_sep", ".")
    skip_labels_list = config.get("skip_labels", [])
    skip_labels_norm = {_norm(lbl) for lbl in skip_labels_list}

    unmatched: list[tuple[str, str]] = []
    for row in rows:
        for sec, zone in zones.items():
            if zone is None:
                continue
            lc, vc, vp = zone
            label = row.get(lc, "")
            if not label:
                continue

            # Skip rows that match skip_labels (e.g., reinsurer company names in financial columns)
            if _norm(label) in skip_labels_norm:
                continue

            raw_cur = row.get(vc, "")
            raw_pri = row.get(vp, "")

            # Both period values may have collapsed into the current column
            if raw_cur and not raw_pri:
                raw_cur, raw_pri = _split_two(raw_cur)

            val_cur = _parse_val(raw_cur, tsep)
            val_pri = _parse_val(raw_pri, tsep)
            if val_cur is None and val_pri is None:
                continue

            tmpl_row = _find_row(label, sec, idx, ovr)
            if tmpl_row is None:
                unmatched.append((sec, label))
                continue

            col_cur, col_pri = _VAL_COLS[sec]
            if val_cur is not None:
                ws[f"{col_cur}{tmpl_row}"] = val_cur
            if val_pri is not None:
                ws[f"{col_pri}{tmpl_row}"] = val_pri

    if unmatched:
        seen: set[tuple[str, str]] = set()
        cid = config.get("company_id", "?")
        print(f"  [WARN] {cid}: {len({k for k in unmatched})} unmatched label(s):")
        for sec, lbl in unmatched:
            if (sec, lbl) not in seen:
                print(f"    {sec}: {lbl!r}")
                seen.add((sec, lbl))

    fill_bottom_section(ws, rows, config)
    wb.save(out_path)
