"""Build an openpyxl workbook from extracted rows + company config + report period."""
from __future__ import annotations

import calendar
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_MONTHS_ID_FULL = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}
_MONTHS_ID_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "Mei", 6: "Jun", 7: "Jul", 8: "Agt",
    9: "Sep", 10: "Okt", 11: "Nov", 12: "Des",
}
_MONTHS_EN_FULL = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}
_MONTHS_EN_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}
_PAGE_SIZES = {"A3": 8, "A4": 9, "LETTER": 1}


def _last_day(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def compute_dates(period: str, config: dict) -> dict:
    """Return a dict of template placeholders derived from the period string (YYYY-MM)."""
    year, month = int(period[:4]), int(period[5:7])
    day = _last_day(year, month)

    cmp = config.get("period", {}).get("comparison", "prior_year_same_month")
    p_year, p_month = (year - 1, 12) if cmp == "prior_year_end" else (year - 1, month)
    p_day = _last_day(p_year, p_month)

    return {
        "day": str(day),
        "month_full": _MONTHS_ID_FULL[month],
        "MONTH_FULL": _MONTHS_ID_FULL[month].upper(),
        "month_abbr": _MONTHS_ID_ABBR[month],
        "MONTH_ABBR": _MONTHS_ID_ABBR[month].upper(),
        "month_en": _MONTHS_EN_FULL[month],
        "MONTH_EN": _MONTHS_EN_FULL[month].upper(),
        "month_en_abbr": _MONTHS_EN_ABBR[month],
        "MONTH_EN_ABBR": _MONTHS_EN_ABBR[month].upper(),
        "YEAR4": str(year),
        "YEAR2": str(year)[-2:],
        "prior_day": str(p_day),
        "prior_month_full": _MONTHS_ID_FULL[p_month],
        "PRIOR_MONTH_FULL": _MONTHS_ID_FULL[p_month].upper(),
        "prior_month_abbr": _MONTHS_ID_ABBR[p_month],
        "PRIOR_MONTH_ABBR": _MONTHS_ID_ABBR[p_month].upper(),
        "prior_month_en": _MONTHS_EN_FULL[p_month],
        "PRIOR_MONTH_EN": _MONTHS_EN_FULL[p_month].upper(),
        "prior_month_en_abbr": _MONTHS_EN_ABBR[p_month],
        "PRIOR_MONTH_EN_ABBR": _MONTHS_EN_ABBR[p_month].upper(),
        "PRIOR_YEAR4": str(p_year),
        "PRIOR_YEAR2": str(p_year)[-2:],
    }


def _fmt(template: str, dates: dict) -> str:
    try:
        return template.format(**dates)
    except (KeyError, ValueError):
        return template


def _apply(cell, *, font=None, fill=None, align=None, border=None) -> None:
    if font   is not None: cell.font      = font
    if fill   is not None: cell.fill      = fill
    if align  is not None: cell.alignment = align
    if border is not None: cell.border    = border


def save_workbook(rows: list[dict], config: dict, period: str, out_path: Path) -> None:
    """Build a workbook from extracted rows + config + period and write it to out_path."""
    dates = compute_dates(period, config)

    wb = Workbook()
    ws = wb.active
    ws.title = config.get("sheet_name", config["company_id"])

    # ── Page setup ──────────────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = config.get("zoom", 70)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = _PAGE_SIZES.get(config.get("page_size", "LETTER"), 1)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    m = config.get("margins", {})
    ws.page_margins.left   = m.get("left",   0.2)
    ws.page_margins.right  = m.get("right",  0.2)
    ws.page_margins.top    = m.get("top",    0.25)
    ws.page_margins.bottom = m.get("bottom", 0.25)
    ws.page_margins.header = m.get("header", 0.1)
    ws.page_margins.footer = m.get("footer", 0.1)

    # ── Column widths ────────────────────────────────────────────────────────────
    for col, w in config.get("col_widths", {}).items():
        ws.column_dimensions[col].width = w

    # ── Style objects ────────────────────────────────────────────────────────────
    clrs = config.get("colors", {})
    fill_section = PatternFill("solid", fgColor=clrs.get("section", "000000"))
    fill_header  = PatternFill("solid", fgColor=clrs.get("header",  "D9D9D9"))
    fill_year    = PatternFill("solid", fgColor=clrs.get("year",    "EFEFEF"))
    fill_subtle  = PatternFill("solid", fgColor=clrs.get("subtle",  "F2F2F2"))
    fill_en      = PatternFill("solid", fgColor=clrs.get("english", "F8F8FF"))

    body_sz    = config.get("body_font_size",    6.5)
    section_sz = config.get("section_font_size", 7.0)
    en_sz      = config.get("english_font_size", 5.8)

    f_section = Font(name="Times New Roman", size=section_sz, bold=True, color="FFFFFF")
    f_table   = Font(name="Times New Roman", size=section_sz, bold=True)
    f_year    = Font(name="Times New Roman", size=section_sz, bold=True)
    f_body    = Font(name="Times New Roman", size=body_sz)
    f_bold    = Font(name="Times New Roman", size=body_sz, bold=True)
    f_en      = Font(name="Times New Roman", size=en_sz, italic=True)

    a_ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    a_rgt = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    thin   = Side(style="thin",   color="000000")
    medium = Side(style="medium", color="000000")
    b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_box  = Border(left=medium, right=medium, top=medium, bottom=medium)

    # ── Header rows 1–6 ──────────────────────────────────────────────────────────
    header_rows = config.get("header_rows", {})
    for row_num in sorted(header_rows.keys(), key=int):
        rc = header_rows[row_num]
        ws.merge_cells(f"A{row_num}:P{row_num}")
        ws[f"A{row_num}"] = _fmt(rc["text"], dates)
        _apply(
            ws[f"A{row_num}"],
            font=Font(name="Times New Roman", size=rc.get("size", 8), bold=rc.get("bold", False)),
            align=a_ctr,
        )
        ws.row_dimensions[int(row_num)].height = rc.get("height", 13)

    # ── Section bars rows 7–9 ────────────────────────────────────────────────────
    sb = config.get("section_bars", {})
    for i, h in enumerate(sb.get("row_heights", [17, 13, 13])):
        ws.row_dimensions[7 + i].height = h
    for item in sb.get("items", []):
        rng = item["range"]
        ws.merge_cells(rng)
        tl = rng.split(":")[0]
        ws[tl] = _fmt(item["text"], dates)
        _apply(ws[tl], font=f_section, fill=fill_section, align=a_ctr, border=b_box)

    # ── Column group headers row 10 ──────────────────────────────────────────────
    ch = config.get("col_headers", {})
    ws.row_dimensions[10].height = ch.get("height", 18)
    for item in ch.get("items", []):
        rng = item["range"]
        ws.merge_cells(rng)
        tl = rng.split(":")[0]
        ws[tl] = item["text"]
        _apply(ws[tl], font=f_table, fill=fill_header, align=a_ctr, border=b_thin)

    # ── Year labels row 11 ───────────────────────────────────────────────────────
    yl = config.get("year_labels", {})
    ws.row_dimensions[11].height = yl.get("height", 15)
    period_cfg = config.get("period", {})
    current_lbl = _fmt(period_cfg.get("current_label", "{YEAR4}"), dates)
    prior_lbl   = _fmt(period_cfg.get("prior_label",   "{PRIOR_YEAR4}"), dates)

    for cell_ref, kind in yl.get("cells", {}).items():
        ws[cell_ref] = current_lbl if kind == "current" else prior_lbl
        _apply(ws[cell_ref], font=f_year, fill=fill_year, align=a_ctr, border=b_thin)

    for cell_ref, text in yl.get("static_cells", {}).items():
        ws[cell_ref] = text
        _apply(
            ws[cell_ref],
            font=Font(name="Times New Roman", size=6, italic=True),
            fill=fill_en, align=a_ctr, border=b_thin,
        )

    # ── Data rows 12+ ────────────────────────────────────────────────────────────
    value_cols   = set(config.get("value_cols", []))
    body_cols    = tuple(config.get("body_cols", []))
    english_cols = set(config.get("english_cols", []))
    h_scale      = config.get("row_height_scale", 1.15)
    h_max        = config.get("row_height_max")

    start_row = 12
    for idx, row in enumerate(rows):
        row_num = start_row + idx
        flags   = row.get("_flags", {})
        gap     = row.get("_gap", 8.0)
        h = gap * h_scale
        if h_max is not None:
            h = min(h_max, h)
        ws.row_dimensions[row_num].height = max(7.0, h)

        for col in body_cols:
            val = row.get(col)
            if not val:
                continue
            ws[f"{col}{row_num}"] = val
            is_val = col in value_cols
            is_en  = col in english_cols
            _apply(
                ws[f"{col}{row_num}"],
                font=(f_bold if flags.get("bold") and not is_en else f_en if is_en else f_body),
                align=(a_rgt if is_val else a_ctr if flags.get("center") else a_lft),
                border=b_thin,
                fill=(fill_en if is_en
                      else fill_subtle if flags.get("bold") and not is_val else None),
            )

    last_row = start_row + len(rows) - 1
    ws.print_area = f"A1:P{last_row}"
    ws.freeze_panes = "A12"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
