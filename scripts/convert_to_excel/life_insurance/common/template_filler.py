"""Fill the OJK standardized life insurance template with values extracted from a company's PDF.

Strategy:
  1. Copy the template file as base (all styling, merged cells preserved).
  2. Fill company header + period in header rows.
  3. For each extracted row, identify which template row the label matches
     and write the numeric values into the correct value columns.
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from copy import copy

TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "laporan_keuangan_template_asuransi_jiwa.xlsx"

_VAL_COLS = {
    "assets":      ("C",  "D"),
    "liabilities": ("G",  "H"),
    "income":      ("K",  "L"),
    "health":      ("O",  "P"),
}

_LABEL_COLS = {
    "assets":      "B",
    "liabilities": "F",
    "income":      "J",
    "health":      "N",
}

_DATA_ROWS = range(12, 62)

_SYNONYMS = {
    "retrosesi": "reasuransi",
    "kewajiban": "liabilitas",
    "hutang":    "utang",
    "piutang":   "tagihan",
    "risikio":   "risiko",
    "reksadana": "reksa dana",
    "netto":     "neto",
}


def _norm(text: str) -> str:
    """Lowercase, strip item-number prefixes + parentheticals, apply synonyms."""
    if not text:
        return ""
    t = str(text).strip()
    t = re.sub(r"^(IV|IX|V?I{1,3}|XI{0,2})\s*\.?\s+", "", t, flags=re.IGNORECASE)
    t = re.sub(r"^\d+\s*\.?\s+", "", t)
    t = re.sub(r"^[a-eA-E]\s*\.?\s+", "", t)
    t = re.sub(r"\s*\([^)]*\)", "", t)
    t = t.replace("*", "")
    t = " ".join(t.split()).lower()
    for src, dst in _SYNONYMS.items():
        t = t.replace(src, dst)
    return t


_NUM_RE = re.compile(r"^\(?-?[\d.,]+\)?%?$")


def _parse_val(text: str, thousands_sep: str = ".") -> float | None:
    """Parse Indonesian or US number format → float."""
    if not text:
        return None
    t = str(text).strip()
    if t in ("-", ""):
        return None
    negative = t.startswith("(") and t.endswith(")")
    if negative:
        t = t[1:-1]
    t = t.rstrip("%").replace(" ", "")
    if not t:
        return None

    decimal_sep = "," if thousands_sep == "." else "."

    if decimal_sep in t:
        parts = t.rsplit(decimal_sep, 1)
        try:
            val = float(parts[0].replace(thousands_sep, "") + "." + parts[1])
        except (ValueError, IndexError):
            return None
    else:
        try:
            val = float(t.replace(thousands_sep, ""))
        except ValueError:
            return None
    return -val if negative else val


def _split_two(s: str) -> tuple[str, str]:
    """Split a string that contains two space-separated numbers."""
    if not s:
        return "", ""
    parts = s.split()
    if len(parts) >= 2 and _NUM_RE.match(parts[0]) and _NUM_RE.match(parts[-1]):
        return parts[0], parts[-1]
    return s, ""


def _build_index(ws) -> dict[str, dict[str, int]]:
    """Build {section → {normalized_label → row_number}} from the template sheet."""
    idx: dict[str, dict[str, int]] = {s: {} for s in _LABEL_COLS}
    for row in _DATA_ROWS:
        for sec, col in _LABEL_COLS.items():
            v = ws[f"{col}{row}"].value
            if v:
                n = _norm(str(v))
                if n and n not in idx[sec]:
                    idx[sec][n] = row
    return idx


def _find_row(
    label: str,
    section: str,
    idx: dict[str, dict[str, int]],
    ovr: dict[str, dict[str, int]],
) -> int | None:
    n = _norm(label)
    if not n:
        return None
    if n in ovr.get(section, {}):
        return ovr[section][n]
    if n in idx[section]:
        return idx[section][n]
    for t, row in idx[section].items():
        if n in t:
            return row
    for t, row in idx[section].items():
        if t in n:
            return row
    return None


def _copy_cell_style(source_cell, target_cell) -> None:
    """Copy formatting from source to target cell."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = copy(source_cell.number_format)


def _get_writable_cell(ws, col: str, row: int):
    """Get the cell to write to, handling merged cells."""
    cell = ws[f"{col}{row}"]
    if cell.__class__.__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= cell.column <= merged_range.max_col and \
               merged_range.min_row <= cell.row <= merged_range.max_row:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _fmt(template: str, dates: dict) -> str:
    try:
        return template.format(**dates)
    except (KeyError, ValueError):
        return template


def compute_dates(period: str, config: dict) -> dict:
    """Return a dict of template placeholders derived from the period string (YYYY-MM)."""
    import calendar
    year, month = int(period[:4]), int(period[5:7])
    day = calendar.monthrange(year, month)[1]

    cmp = config.get("period", {}).get("comparison", "prior_year_same_month")
    p_year, p_month = (year - 1, 12) if cmp == "prior_year_end" else (year - 1, month)
    p_day = calendar.monthrange(p_year, p_month)[1]

    _MONTHS_ID_FULL = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember",
    }
    _MONTHS_ID_ABBR = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
        5: "Mei", 6: "Jun", 7: "Jul", 8: "Agt",
        9: "Sep", 10: "Okt", 11: "Nov", 12: "Des",
    }

    return {
        "day": str(day),
        "month_full": _MONTHS_ID_FULL[month],
        "MONTH_FULL": _MONTHS_ID_FULL[month].upper(),
        "month_abbr": _MONTHS_ID_ABBR[month],
        "YEAR4": str(year),
        "YEAR2": str(year)[-2:],
        "prior_day": str(p_day),
        "prior_month_full": _MONTHS_ID_FULL[p_month],
        "PRIOR_YEAR4": str(p_year),
        "PRIOR_YEAR2": str(p_year)[-2:],
    }


def fill_template(
    rows: list[dict],
    config: dict,
    period: str,
    out_path: Path,
) -> None:
    """Copy template, write company header + extracted numbers, save to out_path."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE_PATH, out_path)

    wb = load_workbook(out_path)
    ws = wb.active
    dates = compute_dates(period, config)

    # Fill header rows
    hr = config.get("header_rows", {})
    for row_num in [1, 2, 3]:
        cell = _get_writable_cell(ws, "A", row_num)
        cell.value = _fmt(hr.get(row_num, {}).get("text", ""), dates)

    # Build label index
    idx = _build_index(ws)
    raw_ovr = config.get("label_map", {})
    ovr: dict[str, dict[str, int]] = {
        sec: {_norm(k): v for k, v in mapping.items()}
        for sec, mapping in raw_ovr.items()
    }

    # Zone config
    defaults = {
        "assets":      ("B", "C", "D"),
        "liabilities": ("F", "G", "H"),
        "income":      ("J", "K", "L"),
        "health":      ("N", "O", "P"),
    }
    zcfg = config.get("zones", {})
    zones: dict[str, tuple[str, str, str] | None] = {}
    for sec, (dl, dc, dp) in defaults.items():
        zs = zcfg.get(sec)
        if zs is False or zs == "false":
            zones[sec] = None
        elif isinstance(zs, dict):
            zones[sec] = (zs.get("label_col", dl),
                          zs.get("val_current", dc),
                          zs.get("val_prior",   dp))
        else:
            zones[sec] = (dl, dc, dp)

    # Fill data rows
    tsep = config.get("thousands_sep", ".")
    skip_labels_list = config.get("skip_labels", [])
    skip_labels_norm = {_norm(lbl) for lbl in skip_labels_list}

    unmatched: list[tuple[str, str]] = []
    for row in rows:
        for sec, zone in zones.items():
            if zone is None:
                continue
            lc, vc, vp = zone
            label = row.get(lc, "")
            if not label:
                continue

            if _norm(label) in skip_labels_norm:
                continue

            raw_cur = row.get(vc, "")
            raw_pri = row.get(vp, "")

            if raw_cur and not raw_pri:
                raw_cur, raw_pri = _split_two(raw_cur)

            val_cur = _parse_val(raw_cur, tsep)
            val_pri = _parse_val(raw_pri, tsep)
            if val_cur is None and val_pri is None:
                continue

            tmpl_row = _find_row(label, sec, idx, ovr)
            if tmpl_row is None:
                unmatched.append((sec, label))
                continue

            col_cur, col_pri = _VAL_COLS[sec]
            if val_cur is not None:
                cell = _get_writable_cell(ws, col_cur, tmpl_row)
                cell.value = val_cur
            if val_pri is not None:
                cell = _get_writable_cell(ws, col_pri, tmpl_row)
                cell.value = val_pri

    if unmatched:
        seen: set[tuple[str, str]] = set()
        cid = config.get("company_id", "?")
        print(f"  [WARN] {cid}: {len({k for k in unmatched})} unmatched label(s):")
        for sec, lbl in unmatched:
            if (sec, lbl) not in seen:
                print(f"    {sec}: {lbl!r}")
                seen.add((sec, lbl))

    wb.save(out_path)
